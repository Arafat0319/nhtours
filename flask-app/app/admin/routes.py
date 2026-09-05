from datetime import date, datetime, timedelta
from html import unescape
import json
import os
import re
from flask import render_template, redirect, url_for, flash, request, jsonify, current_app, send_file, abort
from flask_login import login_user, logout_user, current_user, login_required
from app.admin.decorators import admin_required
from sqlalchemy.orm import joinedload
from sqlalchemy import or_, and_, case, func
from collections import defaultdict
from app import db
from app.admin import bp
from app.admin.forms import LoginForm, TripForm, CityForm, ClientForm, TripBasicsForm, TripDescriptionForm, TripPackagesForm, TripAddonsForm, TripParticipantForm, TripCouponForm, EditBookingForm, TestimonialForm
from app.models import User, Trip, City, Client, Lead, Testimonial, TripPackage, TripAddOn, CustomQuestion, DiscountCode, Booking, BookingParticipant, BookingAddOn, BookingPackage, Payment, Message, InstallmentPayment, PendingBooking
from app.payments import create_checkout_session, booking_package_unit_price, booking_addon_unit_price
from app.utils import (
    save_image,
    send_email_via_ses,
    generate_installment_token,
    generate_export_token,
    verify_export_token,
    render_branded_customer_message,
)
from app.testimonial_data import assign_carousel_sort_order
from app.security_audit import (
    is_login_rate_limited,
    record_login_failure,
    record_login_success,
    record_logout,
)


from app.trip_status import (
    active_booking_count,
    check_trip_completion,
    copy_trip_status,
    get_trip_counts,
    get_trip_publish_gaps,
    gap_labels,
    migrate_legacy_trip_statuses,
    publish_trip,
    resolve_trips_filter,
    sync_trip_lifecycle,
    trip_list_bucket,
    unpublish_trip,
)


@bp.context_processor
def inject_trip_counts_nav():
    """侧栏五栏计数（My Trips 导航）。"""
    try:
        return {'trip_counts': get_trip_counts()}
    except Exception:
        return {}


@bp.context_processor
def inject_trip_builder_publish_gaps():
    """Trip Builder 侧边栏显示尚未满足的发布条件。"""
    if request.endpoint != 'admin.trip_builder':
        return {}
    trip_id = (request.view_args or {}).get('id')
    if not trip_id:
        return {}
    trip = Trip.query.get(trip_id)
    if not trip:
        return {}
    return {
        'publish_gaps': get_trip_publish_gaps(trip),
        'trip_bucket': trip_list_bucket(trip),
        'active_booking_count': active_booking_count(trip) if trip else 0,
    }

def _prefetch_booking_finance_graph(trip_id):
    """一次拉齐行程订单套餐/附加，避免 trips 列表 stats 的 dynamic N+1。"""
    bookings = (
        Booking.query.filter_by(trip_id=trip_id)
        .options(joinedload(Booking.client))
        .all()
    )
    if not bookings:
        return [], {}, {}, {}

    booking_ids = [b.id for b in bookings]
    bp_by_booking = defaultdict(list)
    for bp in (
        BookingPackage.query.filter(BookingPackage.booking_id.in_(booking_ids))
        .options(joinedload(BookingPackage.package))
        .all()
    ):
        bp_by_booking[bp.booking_id].append(bp)

    participants = BookingParticipant.query.filter(
        BookingParticipant.booking_id.in_(booking_ids)
    ).all()
    participant_by_id = {p.id: p for p in participants}

    participant_addons_by_booking = defaultdict(list)
    if participant_by_id:
        for ba in (
            BookingAddOn.query.filter(BookingAddOn.participant_id.in_(list(participant_by_id)))
            .options(joinedload(BookingAddOn.addon))
            .all()
        ):
            p = participant_by_id.get(ba.participant_id)
            if p:
                participant_addons_by_booking[p.booking_id].append(ba)

    booking_addons_by_booking = defaultdict(list)
    for ba in (
        BookingAddOn.query.filter(BookingAddOn.booking_id.in_(booking_ids))
        .options(joinedload(BookingAddOn.addon))
        .all()
    ):
        booking_addons_by_booking[ba.booking_id].append(ba)

    return bookings, bp_by_booking, participant_addons_by_booking, booking_addons_by_booking


def calculate_trip_stats(trip):
    """
    计算行程的统计信息（参与者数量、已付金额、应付金额）
    返回: dict with 'participants_count', 'amount_paid', 'amount_gross', 'amount_discount',
    'amount_expected', 'amount_refunded', 'amount_available'
    
    重要说明：
    - amount_paid: 来自 Booking.amount_paid，是客户实际支付的基础金额（不含 Stripe 手续费）
    - amount_gross: 原价总额（套餐 + 附加项）
    - amount_discount: 折扣总额
    - amount_expected: 净应收金额（gross - discount）
    - amount_refunded: Σ 已退基础额（与 Manage Refunded 同口径）
    - amount_available: Σ balance due（待收；已退款不计入欠款）
    """
    from app.payments import booking_balance_due, booking_refunded_total

    bookings, bp_by_booking, participant_addons_by_booking, booking_addons_by_booking = (
        _prefetch_booking_finance_graph(trip.id)
    )

    participants_count = (
        db.session.query(func.count(BookingParticipant.id))
        .join(Booking, BookingParticipant.booking_id == Booking.id)
        .filter(Booking.trip_id == trip.id)
        .scalar()
    ) or 0

    # 计算已付金额（Booking.amount_paid 只记录基础金额，不含手续费）
    amount_paid = sum(b.amount_paid or 0.0 for b in bookings)
    
    # 计算应付金额（套餐 + 附加项 - 折扣）
    amount_gross = 0.0
    amount_discount = 0.0
    amount_expected = 0.0
    amount_refunded = 0.0
    amount_available = 0.0
    
    for b in bookings:
        booking_gross = 0.0
        has_packages = False
        
        # Calculate expected amount from BookingPackages
        for bp in bp_by_booking.get(b.id, []):
            if bp.package:
                package_price = booking_package_unit_price(bp)
                quantity = int(bp.quantity) if bp.quantity is not None else 1
                booking_gross += package_price * quantity
                has_packages = True
        
        # Add add-ons prices (收集所有 add-ons，避免重复计算)
        seen_addon_ids = set()
        
        for booking_addon in participant_addons_by_booking.get(b.id, []):
            if booking_addon.addon and booking_addon.id not in seen_addon_ids:
                addon_price = booking_addon_unit_price(booking_addon)
                quantity = int(booking_addon.quantity) if booking_addon.quantity is not None else 1
                booking_gross += addon_price * quantity
                seen_addon_ids.add(booking_addon.id)
        
        for booking_addon in booking_addons_by_booking.get(b.id, []):
            if booking_addon.addon and booking_addon.id not in seen_addon_ids:
                addon_price = booking_addon_unit_price(booking_addon)
                quantity = int(booking_addon.quantity) if booking_addon.quantity is not None else 1
                booking_gross += addon_price * quantity
                seen_addon_ids.add(booking_addon.id)
        
        # 折扣金额
        discount = float(b.discount_amount) if b.discount_amount else 0.0
        booking_expected = max(0.0, booking_gross - discount)
        
        # Fallback for legacy bookings without BookingPackages
        if not has_packages:
            booking_gross = float(b.amount_paid) if b.amount_paid is not None else 0.0
            booking_expected = booking_gross
            discount = 0.0
        
        amount_gross += booking_gross
        amount_discount += discount
        amount_expected += booking_expected
        amount_refunded += booking_refunded_total(b)
        due = booking_balance_due(b, expected=booking_expected)
        if due is not None:
            amount_available += due
    
    return {
        'participants_count': participants_count,
        'amount_paid': amount_paid,
        'amount_gross': amount_gross,
        'amount_discount': amount_discount,
        'amount_expected': amount_expected,
        'amount_refunded': round(amount_refunded, 2),
        'amount_available': amount_available
    }

# Force reload


from wtforms import StringField, PasswordField, BooleanField, SubmitField, FloatField, DateField, TextAreaField, SelectMultipleField


@bp.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('admin.trips'))
    
    form = LoginForm()
    if form.validate_on_submit():
        if is_login_rate_limited():
            flash('登录尝试过多，请稍后再试。')
            return redirect(url_for('admin.login'))

        username = (form.username.data or '').strip()
        user = User.query.filter_by(username=username).first()
        if user is None or not user.check_password(form.password.data):
            record_login_failure(username)
            flash('用户名或密码错误')
            return redirect(url_for('admin.login'))
        
        login_user(user, remember=form.remember_me.data)
        record_login_success(user.username)
        next_page = request.args.get('next')
        if not next_page or not next_page.startswith('/'):
            next_page = url_for('admin.trips')
        return redirect(next_page)
        
    return render_template('admin/login.html', title='管理员登录', form=form)

@bp.route('/logout')
def logout():
    if current_user.is_authenticated:
        record_logout(current_user.username)
    logout_user()
    return redirect(url_for('admin.login'))


@bp.route('/')
@bp.route('/dashboard')
@login_required
def dashboard():
    return redirect(url_for('admin.trips'))
    # trip_count = Trip.query.count()
    # client_count = Client.query.count()
    # 暂时没有 Payment 模型或者 Payment 查询逻辑，先设为0
    # payment_total = db.session.query(db.func.sum(Payment.amount)).scalar() or 0
    # 只是简单计数
    
    # return render_template('admin/dashboard.html', title='管理仪表盘', 
    #                        trip_count=trip_count, 
    #                        client_count=client_count)


@bp.route('/trips')
@login_required
def trips():
    from app.utils import pacific_today

    filter_type = resolve_trips_filter(request.args.get('filter', 'future'))
    query = Trip.query
    today = pacific_today()

    # Repair legacy deactivated → unpublished/draft (never auto-publish)
    migrate_legacy_trip_statuses(commit=True)

    if filter_type == 'future':
        query = query.filter(Trip.status == 'published', Trip.start_date > today)
        default_sort = Trip.start_date.asc()
    elif filter_type == 'in_progress':
        query = query.filter(
            Trip.status == 'published',
            Trip.start_date <= today,
            Trip.end_date >= today,
        )
        default_sort = Trip.start_date.asc()
    elif filter_type == 'past':
        query = query.filter(Trip.status == 'published', Trip.end_date < today)
        default_sort = Trip.start_date.desc()
    elif filter_type == 'draft':
        query = query.filter(Trip.status == 'draft')
        default_sort = Trip.updated_at.desc()
    elif filter_type == 'unpublished':
        query = query.filter(Trip.status.in_(('unpublished', 'deactivated')))
        default_sort = Trip.updated_at.desc()
    else:
        default_sort = Trip.created_at.desc()

    search_query = request.args.get('q')
    if search_query:
        query = query.filter(Trip.title.ilike(f'%{search_query}%'))

    sort_by = request.args.get('sort')
    if sort_by == 'created_asc':
        query = query.order_by(Trip.created_at.asc())
    elif sort_by == 'created_desc':
        query = query.order_by(Trip.created_at.desc())
    elif sort_by == 'start_date_asc':
        query = query.order_by(Trip.start_date.asc())
    elif sort_by == 'start_date_desc':
        query = query.order_by(Trip.start_date.desc())
    elif sort_by == 'title_asc':
        query = query.order_by(Trip.title.asc())
    else:
        query = query.order_by(default_sort)

    trips = query.all()
    count = len(trips)
    trip_counts = get_trip_counts()
    trip_stats = {trip.id: calculate_trip_stats(trip) for trip in trips}

    view_type = request.args.get('view', 'list')
    if view_type == 'calendar':
        return render_template('admin/trips/calendar.html', title='行程日历', filter_type=filter_type, count=count, trip_counts=trip_counts)

    return render_template('admin/trips/list_card.html', title='行程管理', trips=trips, filter_type=filter_type, count=count, trip_counts=trip_counts, trip_stats=trip_stats)


@bp.route('/trips/json')
@login_required
def trips_json():
    """返回行程数据供日历使用"""
    trips = Trip.query.filter(Trip.status != 'archived').all()
    events = []
    
    for trip in trips:
        color = trip.color or '#00D1C1' # Default to Trip Color or Cyan
        if trip.status == 'draft':
            color = '#9CA3AF' # Gray overrides custom color for draft status (optional, or let user decide)
            
        events.append({
            'id': trip.id,
            'title': trip.title,
            'start': trip.start_date.isoformat() if trip.start_date else None,
            'end': trip.end_date.isoformat() if trip.end_date else None,
            'url': url_for('admin.edit_trip', id=trip.id),
            'backgroundColor': color,
            'borderColor': color,
            'extendedProps': {
                'spots_sold': trip.spots_sold,
                'capacity': trip.capacity
            }
        })
    
    from flask import jsonify
    return jsonify(events)


@bp.route('/trips/<int:id>/manage')
@login_required
def manage_trip(id):
    trip = Trip.query.get_or_404(id)
    bookings = (
        Booking.query.filter_by(trip_id=trip.id)
        .options(joinedload(Booking.client))
        .all()
    )
    
    # Financial Calculations
    total_gross = 0.0  # 原价总额（未扣除折扣）
    total_discount = 0.0  # 折扣总额
    total_expected = 0.0  # 净应收金额（扣除折扣后）
    booking_balances = {}  # booking_id -> pending balance (None = cancelled / n/a)
    from app.payments import (
        booking_balance_due,
        booking_refunded_total,
        computed_booking_amount_paid,
        ledger_payments_for_booking,
    )

    booking_paid_display = {}
    for b in bookings:
        # 列表 Paid 与弹窗一致：有 Payment 账本时用 Σ(base−refunded)
        if ledger_payments_for_booking(b):
            booking_paid_display[b.id] = computed_booking_amount_paid(b)
        else:
            booking_paid_display[b.id] = round(
                float(b.amount_paid) if b.amount_paid is not None else 0.0, 2
            )

        booking_gross = 0.0  # 该订单原价
        has_packages = False
        seen_addon_ids = set()
        
        # Calculate expected amount from BookingPackages
        for bp in b.booking_packages:
            if bp.package:  # Check if package exists
                package_price = booking_package_unit_price(bp)
                quantity = int(bp.quantity) if bp.quantity is not None else 1
                booking_gross += package_price * quantity
                has_packages = True
        
        # Add add-ons prices (通过 participant.addons)
        for participant in b.participants:
            for booking_addon in participant.addons:
                if booking_addon.addon and booking_addon.id not in seen_addon_ids:
                    addon_price = booking_addon_unit_price(booking_addon)
                    quantity = int(booking_addon.quantity) if booking_addon.quantity is not None else 0
                    booking_gross += addon_price * quantity
                    seen_addon_ids.add(booking_addon.id)
        
        # Add add-ons prices (直接通过 booking.addons)
        for booking_addon in b.addons:
            if booking_addon.addon and booking_addon.id not in seen_addon_ids:
                addon_price = booking_addon_unit_price(booking_addon)
                quantity = int(booking_addon.quantity) if booking_addon.quantity is not None else 0
                booking_gross += addon_price * quantity
                seen_addon_ids.add(booking_addon.id)
        
        # 折扣金额
        discount = float(b.discount_amount) if b.discount_amount else 0.0
        booking_expected = max(0.0, booking_gross - discount)
        
        # Fallback for legacy bookings without BookingPackages
        if not has_packages:
            booking_gross = float(b.amount_paid) if b.amount_paid is not None else 0.0
            booking_expected = booking_gross  # Legacy bookings don't have discount tracking
            discount = 0.0
        
        total_gross += booking_gross
        total_discount += discount
        total_expected += booking_expected

        # 退款不计入欠款：due = expected − paid − refunded
        booking_balances[b.id] = booking_balance_due(b, expected=booking_expected)

    total_paid = round(sum(booking_paid_display.values()), 2)
    total_pending = round(sum(v for v in booking_balances.values() if v is not None), 2)
    total_refunded = round(sum(booking_refunded_total(b) for b in bookings), 2)

    # Calculate add-ons summary for each booking
    booking_addons_summary = {}
    for booking in bookings:
        addons_map = {}
        # 方法1：通过 participant.addons 获取（关联到特定参与者的 addon）
        for participant in booking.participants:
            for booking_addon in participant.addons:
                if booking_addon.addon:  # 确保 addon 存在
                    addon_name = booking_addon.addon.name
                    if addon_name in addons_map:
                        addons_map[addon_name] += booking_addon.quantity
                    else:
                        addons_map[addon_name] = booking_addon.quantity
        
        # 方法2：通过 booking.addons 获取（没有关联到特定参与者的 addon，或 participant_id 为 null）
        for booking_addon in booking.addons:
            if booking_addon.addon and (booking_addon.participant_id is None or booking_addon.participant_id not in [p.id for p in booking.participants]):
                addon_name = booking_addon.addon.name
                if addon_name in addons_map:
                    addons_map[addon_name] += booking_addon.quantity
                else:
                    addons_map[addon_name] = booking_addon.quantity
        
        booking_addons_summary[booking.id] = addons_map
    
    # Booking buyers for New Message (not participants)
    from app.messaging import collect_message_buyers, forced_reply_to_email, recipient_counts_for_trip
    from app.payments import (
        booking_payment_display_status,
        booking_payment_type_display,
        booking_ids_with_payoff,
    )
    message_buyers = collect_message_buyers(trip)
    buyer_count = len(message_buyers)
    message_recipient_counts = recipient_counts_for_trip(trip)
    balance_due_buyer_count = message_recipient_counts.get('payment_due', 0)

    booking_payment_statuses = {
        b.id: booking_payment_display_status(b) for b in bookings
    }
    booking_payment_types = {
        b.id: booking_payment_type_display(b) for b in bookings
    }
    booking_payoff_ids = booking_ids_with_payoff([b.id for b in bookings])

    # Collect all participants with their booking and package info
    all_participants = []
    total_participants_count = 0
    for booking in bookings:
        for participant in booking.participants:
            # Participants Tab：仅本人 Add-ons（订单级见 Bookings）
            participant_addons = []
            seen_addon_ids = set()
            for booking_addon in participant.addons:
                if booking_addon.addon and booking_addon.id not in seen_addon_ids:
                    participant_addons.append({
                        'name': booking_addon.addon.name,
                        'quantity': booking_addon.quantity,
                        'price': booking_addon_unit_price(booking_addon)
                    })
                    seen_addon_ids.add(booking_addon.id)
            for booking_addon in booking.addons:
                if (
                    booking_addon.addon
                    and booking_addon.participant_id == participant.id
                    and booking_addon.id not in seen_addon_ids
                ):
                    participant_addons.append({
                        'name': booking_addon.addon.name,
                        'quantity': booking_addon.quantity,
                        'price': booking_addon_unit_price(booking_addon)
                    })
                    seen_addon_ids.add(booking_addon.id)

            first_name = (getattr(participant, 'first_name', None) or '').strip()
            last_name = (getattr(participant, 'last_name', None) or '').strip()
            if not first_name and not last_name:
                name_parts = (participant.name or '').strip().split(None, 1)
                first_name = name_parts[0] if name_parts else ''
                last_name = name_parts[1] if len(name_parts) > 1 else ''

            buyer_name = (booking.buyer_name or '').strip()
            if not buyer_name and booking.client:
                buyer_name = (booking.client.name or '').strip()
            buyer_email = (booking.get_buyer_email() or '').strip()

            all_participants.append({
                'id': participant.id,
                'name': participant.name,
                'first_name': first_name,
                'last_name': last_name,
                'email': participant.email,
                'phone': participant.phone,
                'gender': getattr(participant, 'gender', None) or '',
                'dob': participant.dob.strftime('%Y-%m-%d') if getattr(participant, 'dob', None) else '',
                'registration_type': getattr(participant, 'registration_type', None) or '',
                'booking_id': booking.id,
                'order_number': booking.order_number or f'#{booking.id}',
                'buyer_name': buyer_name,
                'buyer_email': buyer_email,
                'addons': participant_addons,
                'question_answers': participant.question_answers or {},
                'status': getattr(participant, 'status', None) or 'active',
                'is_withdrawn': (getattr(participant, 'status', None) or 'active') == 'withdrawn',
            })
            total_participants_count += 1

    # Going = 未取消订单上仍 active 的参与者（withdrawn 不计）
    going_count = 0
    for b in bookings:
        if (b.status or '') == 'cancelled':
            continue
        for p in b.participants:
            if (getattr(p, 'status', None) or 'active') != 'withdrawn':
                going_count += 1
    
    # Prepare Add Participant Form (deprecated - now using JSON API via multi-step modal)
    from app.admin.forms import AdminBookingForm
    form = AdminBookingForm()
    # Note: package_id and passenger_count removed - now handled via BookingPackage model
    # Multiple packages can be selected in the multi-step modal
    
    # Get custom questions from trip builder
    custom_questions = trip.questions.all() if trip.questions else []
    
    # Get messages for this trip
    sent_messages = Message.query.filter_by(trip_id=trip.id, status='sent').order_by(Message.sent_at.desc()).all()
    scheduled_messages = Message.query.filter_by(trip_id=trip.id, status='scheduled').order_by(Message.scheduled_at.asc()).all()
    draft_messages = Message.query.filter_by(trip_id=trip.id, status='draft').order_by(Message.created_at.desc()).all()
    
    # Get sender email from config (SES verified From address)
    sender_email = current_app.config.get('SENDER_EMAIL', 'nhtours-noreply@nhtours.com')
    # Display name in customer inbox (not admin username)
    sender_name = current_app.config.get('SENDER_DISPLAY_NAME', 'Nexus Horizons Tours')
    # Messages Reply-To：工作邮箱（与 RECIPIENT_EMAIL 解耦）
    reply_to_default = forced_reply_to_email()
    
    # Calculate trip counts for sidebar navigation
    trip_counts = get_trip_counts()
    
    return render_template('admin/trips/manage.html', 
                           title=f'Manage {trip.title}', 
                           trip=trip, 
                           bookings=bookings,
                           booking_addons_summary=booking_addons_summary,
                           booking_balances=booking_balances,
                           booking_paid_display=booking_paid_display,
                           booking_payment_statuses=booking_payment_statuses,
                           booking_payment_types=booking_payment_types,
                           booking_payoff_ids=booking_payoff_ids,
                           all_participants=all_participants,
                           total_participants_count=total_participants_count,
                           going_count=going_count,
                           message_buyers=message_buyers,
                           buyer_count=buyer_count,
                           balance_due_buyer_count=balance_due_buyer_count,
                           message_recipient_counts=message_recipient_counts,
                           custom_questions=custom_questions,
                           total_gross=total_gross,
                           total_discount=total_discount,
                           total_expected=total_expected,
                           total_paid=total_paid,
                           total_refunded=total_refunded,
                           total_pending=total_pending,
                           sent_messages=sent_messages,
                           scheduled_messages=scheduled_messages,
                           draft_messages=draft_messages,
                           sender_email=sender_email,
                           sender_name=sender_name,
                           reply_to_default=reply_to_default,
                           trip_counts=trip_counts,
                           form=form)



@bp.route('/trips/<int:id>/teacher-view/reset', methods=['POST'])
@login_required
def reset_teacher_view(id):
    """换新老师只读链接；旧 teacher_view_slug 立即失效。"""
    trip = Trip.query.get_or_404(id)
    from app.trip_roster import reset_teacher_view_slug

    slug = reset_teacher_view_slug(trip)
    db.session.commit()
    url = url_for('main.teacher_trip_roster', teacher_view_slug=slug, _external=True)
    return jsonify({'ok': True, 'slug': slug, 'url': url})


@bp.route('/trips/new')
@login_required
def new_trip():
    # Create a draft trip immediately
    trip = Trip(status='draft', title='Untitled Trip')
    db.session.add(trip)
    db.session.commit()
    # Redirect to builder step 1
    return redirect(url_for('admin.trip_builder', id=trip.id, step='basics'))

@bp.route('/trips/<int:id>/edit')
@login_required
def edit_trip(id):
    # Redirect to builder (default to basics)
    return redirect(url_for('admin.trip_builder', id=id, step='basics'))

@bp.route('/trips/<int:id>/builder/<step>', methods=['GET', 'POST'])
@login_required
def trip_builder(id, step):
    trip = Trip.query.get_or_404(id)
    # Get trip counts for sidebar navigation
    trip_counts = get_trip_counts()
    
    if step == 'basics':
        form = TripBasicsForm(obj=trip)
        # Autosave: text fields only (no image), prevents data loss on navigate
        if request.method == 'POST' and request.headers.get('X-Autosave') == '1':
            trip.title = request.form.get('title') or trip.title or ''
            trip.slug = request.form.get('slug') or trip.slug or ''
            trip.destination_text = request.form.get('destination_text') or trip.destination_text or ''
            from app.order_numbers import apply_trip_abbr_from_input, reset_trip_abbr
            from app.trip_roster import apply_teacher_view_slug
            # Title first so reset/suggest uses the latest title
            if request.form.get('trip_abbr_reset') == '1':
                reset_trip_abbr(trip)
            else:
                apply_trip_abbr_from_input(
                    trip, request.form.get('trip_abbr'), suggest_if_missing=True
                )
            try:
                apply_teacher_view_slug(trip, request.form.get('teacher_view_slug'))
            except ValueError as e:
                return jsonify({'success': False, 'error': str(e)}), 400
            for f in ('start_date', 'end_date', 'registration_date'):
                v = request.form.get(f)
                if v:
                    try:
                        setattr(trip, f, datetime.strptime(v, '%Y-%m-%d').date() if isinstance(v, str) else v)
                    except (ValueError, TypeError):
                        pass
            try:
                cap = request.form.get('capacity')
                trip.capacity = int(cap) if cap not in (None, '') else trip.capacity
            except (ValueError, TypeError):
                pass
            try:
                min_cap = request.form.get('min_capacity')
                trip.min_capacity = int(min_cap) if min_cap not in (None, '') else trip.min_capacity
            except (ValueError, TypeError):
                pass
            if request.form.get('color'):
                trip.color = request.form.get('color')
            db.session.commit()
            check_trip_completion(trip)
            return jsonify({
                'success': True,
                'status': trip.status,
                'publish_gaps': get_trip_publish_gaps(trip),
                'trip_abbr': trip.trip_abbr,
                'teacher_view_slug': trip.teacher_view_slug,
            })

        if form.validate_on_submit():
            # Handle explicit remove (优先于上传)
            if request.form.get('remove_hero_image') == '1' and trip.hero_image:
                old_path = os.path.join(current_app.static_folder, trip.hero_image)
                if os.path.isfile(old_path):
                    try:
                        os.remove(old_path)
                    except OSError:
                        pass
                trip.hero_image = None
            # Handle Image Upload
            elif form.hero_image.data and hasattr(form.hero_image.data, 'filename') and form.hero_image.data.filename:
                try:
                    image_path = save_image(form.hero_image.data, folder='trip_images')
                    if image_path:
                        # 新图保存成功后再删除旧图，避免 save_image 失败时丢失原图
                        if trip.hero_image and trip.hero_image != image_path:
                            old_path = os.path.join(current_app.static_folder, trip.hero_image)
                            if os.path.isfile(old_path):
                                try:
                                    os.remove(old_path)
                                except OSError:
                                    pass  # 删除失败则忽略
                        trip.hero_image = image_path
                except ValueError as e:
                    flash(str(e), 'error')
                    return render_template('admin/trips/builder/step_basics.html', title='Trip Basics', trip=trip, form=form, current_step='basics', min_date=date.today().isoformat(), trip_counts=trip_counts, teacher_view_url=url_for('main.teacher_trip_roster', teacher_view_slug=trip.teacher_view_slug, _external=True) if trip.teacher_view_slug else None)
            
            # Manual Populate to avoid overwriting image with FileStorage or issues
            trip.title = form.title.data
            trip.slug = form.slug.data
            trip.destination_text = form.destination_text.data
            trip.start_date = form.start_date.data
            trip.end_date = form.end_date.data
            trip.registration_date = form.registration_date.data
            trip.capacity = form.capacity.data
            trip.min_capacity = form.min_capacity.data
            trip.color = form.color.data

            from app.order_numbers import apply_trip_abbr_from_input, reset_trip_abbr
            from app.trip_roster import apply_teacher_view_slug
            if request.form.get('trip_abbr_reset') == '1':
                reset_trip_abbr(trip)
            else:
                apply_trip_abbr_from_input(
                    trip,
                    form.trip_abbr.data,
                    suggest_if_missing=True,
                )
            try:
                apply_teacher_view_slug(trip, form.teacher_view_slug.data)
            except ValueError as e:
                flash(str(e), 'error')
                form.teacher_view_slug.errors.append(str(e))
                return render_template(
                    'admin/trips/builder/step_basics.html',
                    title='Trip Basics',
                    trip=trip,
                    form=form,
                    current_step='basics',
                    min_date=date.today().isoformat(),
                    trip_counts=trip_counts,
                    teacher_view_url=url_for(
                        'main.teacher_trip_roster',
                        teacher_view_slug=trip.teacher_view_slug,
                        _external=True,
                    ) if trip.teacher_view_slug else None,
                )
            
            db.session.commit()
            # Check if trip is complete and update status
            check_trip_completion(trip)
            # 仅删除封面时留在本页，其他提交跳转到下一步
            if request.form.get('remove_hero_image') == '1':
                flash('Cover photo removed.', 'success')
                return redirect(url_for('admin.trip_builder', id=trip.id, step='basics'))
            return redirect(url_for('admin.trip_builder', id=trip.id, step='description'))
        # Prefill abbr: input holds letters only; YYMM shown as read-only prefix in UI
        abbr_follows_title = True
        from app.order_numbers import sanitize_stored_trip_abbr, trip_abbr_follows_title
        from app.trip_roster import ensure_teacher_view_slug
        letters = sanitize_stored_trip_abbr(trip)
        if letters != (trip.trip_abbr or ''):
            trip.trip_abbr = letters
        ensure_teacher_view_slug(trip)
        if db.session.is_modified(trip):
            db.session.commit()
        if request.method == 'GET':
            form.trip_abbr.data = letters
            form.teacher_view_slug.data = trip.teacher_view_slug or ''
        abbr_follows_title = trip_abbr_follows_title(trip)
        teacher_view_url = None
        if trip.teacher_view_slug:
            teacher_view_url = url_for(
                'main.teacher_trip_roster',
                teacher_view_slug=trip.teacher_view_slug,
                _external=True,
            )
        return render_template(
            'admin/trips/builder/step_basics.html',
            title='Trip Basics',
            trip=trip,
            form=form,
            current_step='basics',
            min_date=date.today().isoformat(),
            trip_counts=trip_counts,
            abbr_follows_title=abbr_follows_title,
            teacher_view_url=teacher_view_url,
        )
    
    elif step == 'description':
        form = TripDescriptionForm(obj=trip)
        # Always expose includes/excludes as JSON strings for the builder JS
        if request.method == 'GET':
            form.trip_includes.data = json.dumps(trip.trip_includes or [])
            form.trip_excludes.data = json.dumps(trip.trip_excludes or [])

        # Autosave: XHR request with X-Autosave header, save without full validation
        if request.method == 'POST' and request.headers.get('X-Autosave') == '1':
            desc = request.form.get('description', '')
            includes_raw = request.form.get('trip_includes', '[]') or '[]'
            excludes_raw = request.form.get('trip_excludes', '[]') or '[]'
            try:
                trip.description = desc
                trip.trip_includes = json.loads(includes_raw)
                trip.trip_excludes = json.loads(excludes_raw)
                if not isinstance(trip.trip_includes, list):
                    trip.trip_includes = []
                if not isinstance(trip.trip_excludes, list):
                    trip.trip_excludes = []
                db.session.commit()
                check_trip_completion(trip)
                return jsonify({'success': True, 'status': trip.status, 'publish_gaps': get_trip_publish_gaps(trip)})
            except (ValueError, TypeError) as e:
                return jsonify({'success': False, 'error': str(e) or 'Invalid JSON'}), 400

        if form.validate_on_submit():
            trip.description = form.description.data
            try:
                includes_raw = form.trip_includes.data or '[]'
                excludes_raw = form.trip_excludes.data or '[]'
                trip.trip_includes = json.loads(includes_raw)
                trip.trip_excludes = json.loads(excludes_raw)
                if not isinstance(trip.trip_includes, list):
                    trip.trip_includes = []
                if not isinstance(trip.trip_excludes, list):
                    trip.trip_excludes = []
            except (ValueError, TypeError):
                flash("Invalid JSON data for includes/excludes", "error")
                return render_template('admin/trips/builder/step_description.html', title='Trip Description', trip=trip, form=form, current_step='description', trip_counts=trip_counts)

            db.session.commit()
            check_trip_completion(trip)
            return redirect(url_for('admin.trip_builder', id=trip.id, step='packages'))

        if request.method == 'POST':
            # Surface why Save & Continue did not advance (CSRF / field errors)
            for field_name, errors in form.errors.items():
                for err in errors:
                    flash(f"{field_name}: {err}", "error")

        return render_template('admin/trips/builder/step_description.html', title='Trip Description', trip=trip, form=form, current_step='description', trip_counts=trip_counts)

    elif step == 'packages':
        form = TripPackagesForm()
        if request.method == 'GET':
            packages = trip.packages.all()
            packages_list = []
            for p in packages:
                pkg_dict = {
                    'id': p.id,
                    'name': p.name,
                    'description': p.description,
                    'price': p.price,
                    'capacity': p.capacity,
                    'status': p.status,
                    'payment_plan_config': p.payment_plan_config,
                    'currency': p.currency
                }
                packages_list.append(pkg_dict)
            form.packages_json.data = json.dumps(packages_list)

        if form.validate_on_submit():
            try:
                packages_data = json.loads(form.packages_json.data) if form.packages_json.data else []
                # Validate payment plan amounts: deposit + installments must equal price
                for pkg_data in packages_data:
                    ppc = pkg_data.get('payment_plan_config') or {}
                    if not ppc.get('enabled'):
                        continue
                    price = float(pkg_data.get('price', 0) or 0)
                    deposit = float(ppc.get('deposit_amount', 0) or 0)
                    installments = ppc.get('installments') or []
                    inst_total = sum(float(i.get('amount', 0) or 0) for i in installments)
                    plan_total = deposit + inst_total
                    if abs(price - plan_total) > 0.01:
                        name = pkg_data.get('name', 'Package')
                        flash(f"Package '{name}': payment plan total does not match price.", 'error')
                        return render_template('admin/trips/builder/step_packages.html', title='Packages', trip=trip, form=form, current_step='packages', trip_counts=trip_counts)

                current_packages = {p.id: p for p in trip.packages}
                processed_ids = []

                for pkg_data in packages_data:
                    pkg_id = pkg_data.get('id')
                    
                    booking_deadline_str = pkg_data.get('booking_deadline')
                    booking_deadline = None
                    if booking_deadline_str:
                         try:
                             # Handle potential 'Z' or other formats if needed, but ISO from JS is usually sufficient
                             # Python 3.7+ fromisoformat matches Javascript toISOString() usually
                             booking_deadline = datetime.fromisoformat(booking_deadline_str.replace('Z', '+00:00'))
                         except ValueError:
                             pass

                    if pkg_id and pkg_id in current_packages:
                        # Update existing
                        pkg = current_packages[pkg_id]
                        pkg.name = pkg_data.get('name')
                        pkg.description = pkg_data.get('description')
                        pkg.price = float(pkg_data.get('price', 0))
                        cap = pkg_data.get('capacity')
                        pkg.capacity = int(cap) if (cap is not None and cap != '') else None
                        pkg.payment_plan_config = pkg_data.get('payment_plan_config')
                        pkg.currency = pkg_data.get('currency', 'USD')
                        
                        pkg.booking_deadline = booking_deadline
                        # min_per_booking and max_per_booking removed - customers can add any number of packages
                        
                        processed_ids.append(pkg_id)
                    else:
                        # Create new
                        new_pkg = TripPackage(
                            trip_id=trip.id,
                            name=pkg_data.get('name'),
                            description=pkg_data.get('description'),
                            price=float(pkg_data.get('price', 0)),
                            capacity=int(pkg_data.get('capacity')) if (pkg_data.get('capacity') is not None and pkg_data.get('capacity') != '') else None,
                            payment_plan_config=pkg_data.get('payment_plan_config'),
                            currency=pkg_data.get('currency', 'USD'),
                            
                            booking_deadline=booking_deadline,
                            # min_per_booking and max_per_booking removed
                            
                            status='available'
                        )
                        db.session.add(new_pkg)
                
                # Delete removed packages (block if any booking still references them)
                blocked = []
                for pid, pkg in current_packages.items():
                    if pid not in processed_ids:
                        ref_count = BookingPackage.query.filter_by(package_id=pid).count()
                        if ref_count:
                            blocked.append((pkg.name or f'#{pid}', ref_count))
                        else:
                            db.session.delete(pkg)

                if blocked:
                    db.session.rollback()
                    details = '; '.join(
                        f'"{name}" ({n} booking line{"s" if n != 1 else ""})'
                        for name, n in blocked
                    )
                    flash(
                        f'Cannot delete package(s) still used by bookings: {details}. '
                        f'Edit in place, mark unavailable, or remove those bookings first.',
                        'error',
                    )
                    return render_template(
                        'admin/trips/builder/step_packages.html',
                        title='Packages',
                        trip=trip,
                        form=form,
                        current_step='packages',
                        trip_counts=trip_counts,
                    )

                db.session.commit()
                # Check if trip is complete and update status
                if not check_trip_completion(trip):
                    gaps = get_trip_publish_gaps(trip)
                    missing = ', '.join(gap_labels(gaps))
                    flash(
                        f'Trip saved as Draft. Still needed before it is ready to publish: {missing}.',
                        'warning'
                    )
                else:
                    db.session.refresh(trip)
                    if trip.status == 'unpublished':
                        flash(
                            'Trip is ready. Find it under Unpublished Trips and click Publish when you want it live.',
                            'success',
                        )
                return redirect(url_for('admin.trip_builder', id=trip.id, step='addons'))

            except ValueError as e:
                flash(f"Invalid JSON data for packages: {str(e)}", "error")
                return render_template('admin/trips/builder/step_packages.html', title='Packages', trip=trip, form=form, current_step='packages', trip_counts=trip_counts)

        return render_template('admin/trips/builder/step_packages.html', title='Packages', trip=trip, form=form, current_step='packages', trip_counts=trip_counts)

    elif step == 'addons':
        form = TripAddonsForm()
        if request.method == 'GET':
            addons = trip.add_ons.all()
            addons_list = []
            for a in addons:
                addons_list.append({
                    'id': a.id,
                    'name': a.name,
                    'price': a.price,
                    'description': a.description
                })
            form.addons_json.data = json.dumps(addons_list)
        
        if form.validate_on_submit():
            try:
                addons_data = json.loads(form.addons_json.data) if form.addons_json.data else []
                current_addons = {a.id: a for a in trip.add_ons}
                processed_ids = []

                for addon_data in addons_data:
                    addon_id = addon_data.get('id')
                    try:
                        price_val = float(addon_data.get('price', 0))
                    except (ValueError, TypeError):
                         price_val = 0.0

                    if addon_id and addon_id in current_addons:
                        # Update
                        addon = current_addons[addon_id]
                        addon.name = addon_data.get('name')
                        addon.price = price_val
                        addon.description = addon_data.get('description')
                        processed_ids.append(addon_id)
                    else:
                        # Create
                        new_addon = TripAddOn(
                            trip_id=trip.id,
                            name=addon_data.get('name'),
                            price=price_val,
                            description=addon_data.get('description')
                        )
                        db.session.add(new_addon)
                
                # Delete removed
                for aid, addon in current_addons.items():
                    if aid not in processed_ids:
                        db.session.delete(addon)
                
                db.session.commit()
                # Check if trip is complete and update status
                check_trip_completion(trip)
                return redirect(url_for('admin.trip_builder', id=trip.id, step='buyer_info'))

            except ValueError as e:
                flash(f"Invalid JSON data for addons: {str(e)}", "error")
                return render_template('admin/trips/builder/step_addons.html', title='Add-ons', trip=trip, form=form, current_step='addons', trip_counts=trip_counts)

        return render_template('admin/trips/builder/step_addons.html', title='Add-ons', trip=trip, form=form, current_step='addons', trip_counts=trip_counts)

    elif step == 'buyer_info':
        from app.admin.forms import TripBuyerInfoForm
        from app.models import BuyerInfoField
        form = TripBuyerInfoForm()
        if request.method == 'GET':
            fields = trip.buyer_info_fields.order_by('display_order').all()
            
            # 如果没有配置任何字段，自动创建默认必填字段
            if not fields:
                default_fields = [
                    {'field_name': 'First Name', 'field_type': 'text', 'is_required': True, 'display_order': 0},
                    {'field_name': 'Last Name', 'field_type': 'text', 'is_required': True, 'display_order': 1},
                    {'field_name': 'Email', 'field_type': 'email', 'is_required': True, 'display_order': 2},
                    {'field_name': 'Phone', 'field_type': 'phone', 'is_required': True, 'display_order': 3}
                ]
                for df in default_fields:
                    new_field = BuyerInfoField(
                        trip_id=trip.id,
                        field_name=df['field_name'],
                        field_type=df['field_type'],
                        is_required=df['is_required'],
                        display_order=df['display_order']
                    )
                    db.session.add(new_field)
                db.session.commit()
                fields = trip.buyer_info_fields.order_by('display_order').all()
            
            f_list = []
            for f in fields:
                f_list.append({
                    'id': f.id,
                    'field_name': f.field_name,
                    'field_type': f.field_type,
                    'options': f.options,
                    'is_required': f.is_required,
                    'display_order': f.display_order
                })
            form.fields_json.data = json.dumps(f_list)
        
        if form.validate_on_submit():
            try:
                f_data_list = json.loads(form.fields_json.data) if form.fields_json.data else []
                current_fields = {f.id: f for f in trip.buyer_info_fields}
                processed_ids = []

                for f_item in f_data_list:
                    f_id = f_item.get('id')
                    
                    # Handle options: enforce list or None
                    opts = f_item.get('options')
                    if opts and isinstance(opts, str):
                        opts = [x.strip() for x in opts.split(',')]
                    elif not isinstance(opts, list):
                        opts = None

                    if f_id and f_id in current_fields:
                        # Update
                        f = current_fields[f_id]
                        f.field_name = f_item.get('field_name')
                        f.field_type = f_item.get('field_type')
                        f.options = opts
                        f.is_required = bool(f_item.get('is_required'))
                        f.display_order = f_item.get('display_order', 0)
                        processed_ids.append(f_id)
                    else:
                        # Create
                        new_f = BuyerInfoField(
                            trip_id=trip.id,
                            field_name=f_item.get('field_name'),
                            field_type=f_item.get('field_type'),
                            options=opts,
                            is_required=bool(f_item.get('is_required')),
                            display_order=f_item.get('display_order', 0)
                        )
                        db.session.add(new_f)
                
                # Delete removed
                for fid, f in current_fields.items():
                    if fid not in processed_ids:
                        db.session.delete(f)
                
                db.session.commit()
                # Check if trip is complete and update status
                check_trip_completion(trip)
                return redirect(url_for('admin.trip_builder', id=trip.id, step='participants'))

            except ValueError as e:
                flash(f"Invalid JSON data for fields: {str(e)}", "error")
                return render_template('admin/trips/builder/step_buyer_info.html', title='Buyer Info', trip=trip, form=form, current_step='buyer_info', trip_counts=trip_counts)

        return render_template('admin/trips/builder/step_buyer_info.html', title='Buyer Info', trip=trip, form=form, current_step='buyer_info', trip_counts=trip_counts)

    elif step == 'participants':
        form = TripParticipantForm()
        if request.method == 'GET':
            form.lock_date.data = trip.participants_request_lock_date
            questions = trip.questions.all()
            q_list = []
            for q in questions:
                q_list.append({
                    'id': q.id,
                    'label': q.label,
                    'type': q.type,
                    'options': q.options,
                    'required': q.required
                })
            form.questions_json.data = json.dumps(q_list)
        
        if form.validate_on_submit():
            try:
                # Save Lock Date
                trip.participants_request_lock_date = form.lock_date.data
                
                q_data_list = json.loads(form.questions_json.data) if form.questions_json.data else []
                current_qs = {q.id: q for q in trip.questions}
                processed_ids = []

                for q_item in q_data_list:
                    q_id = q_item.get('id')
                    
                    # Handle options: enforce list or None
                    opts = q_item.get('options')
                    if opts and isinstance(opts, str):
                         # If it came as comma string or anything, split it, though frontend should send array
                         opts = [x.strip() for x in opts.split(',')]
                    elif not isinstance(opts, list):
                        opts = None

                    if q_id and q_id in current_qs:
                        # Update existing question
                        q = current_qs[q_id]
                        q.label = q_item.get('label')
                        q.type = q_item.get('type')
                        q.options = opts
                        q.required = bool(q_item.get('required'))
                        processed_ids.append(q_id)
                    else:
                        # Create new question (q_id is None or doesn't exist)
                        new_q = CustomQuestion(
                            trip_id=trip.id,
                            label=q_item.get('label'),
                            type=q_item.get('type'),
                            options=opts,
                            required=bool(q_item.get('required'))
                        )
                        db.session.add(new_q)
                        db.session.flush()  # 确保获取到新创建的id
                        processed_ids.append(new_q.id)
                
                # Delete removed
                for qid, q in current_qs.items():
                    if qid not in processed_ids:
                        db.session.delete(q)
                
                db.session.commit()
                # Check if trip is complete and update status
                check_trip_completion(trip)
                return redirect(url_for('admin.trip_builder', id=trip.id, step='coupons'))

            except ValueError as e:
                flash(f"Invalid JSON data for questions: {str(e)}", "error")
                return render_template('admin/trips/builder/step_participants.html', title='Participant Info', trip=trip, form=form, current_step='participants', trip_counts=trip_counts)

        return render_template('admin/trips/builder/step_participants.html', title='Participant Info', trip=trip, form=form, current_step='participants', trip_counts=trip_counts)

    elif step == 'coupons':
        form = TripCouponForm()
        if request.method == 'GET':
            codes = trip.discount_codes.all()
            c_list = []
            for c in codes:
                c_list.append({
                    'id': c.id,
                    'code': c.code,
                    'type': c.type,
                    'amount': c.amount
                })
            form.coupons_json.data = json.dumps(c_list)
        
        if form.validate_on_submit():
            try:
                c_data_list = json.loads(form.coupons_json.data) if form.coupons_json.data else []
                current_codes = {c.id: c for c in trip.discount_codes}
                processed_ids = []

                for c_item in c_data_list:
                    c_id = c_item.get('id')
                    
                    if c_id and c_id in current_codes:
                        # Update
                        code = current_codes[c_id]
                        code.code = c_item.get('code').upper()
                        code.type = c_item.get('type')
                        code.amount = float(c_item.get('amount'))
                        processed_ids.append(c_id)
                    else:
                        # Create
                        new_code = DiscountCode(
                            trip_id=trip.id,
                            code=c_item.get('code').upper(),
                            type=c_item.get('type'),
                            amount=float(c_item.get('amount'))
                        )
                        db.session.add(new_code)
                
                # Delete removed
                for cid, code in current_codes.items():
                    if cid not in processed_ids:
                        db.session.delete(code)
                
                db.session.commit()
                # Check if trip is complete and update status
                check_trip_completion(trip)
                
                # Refresh trip from database to get updated status
                db.session.refresh(trip)
                
                flash('Trip configuration saved successfully!', 'success')
                
                return redirect(url_for('admin.trips', filter=trip_list_bucket(trip)))
            
            except ValueError as e:
                flash(f"Invalid JSON data for coupons: {str(e)}", "error")
                return render_template('admin/trips/builder/step_coupons.html', title='Coupons', trip=trip, form=form, current_step='coupons', trip_counts=trip_counts)

        return render_template('admin/trips/builder/step_coupons.html', title='Coupons', trip=trip, form=form, current_step='coupons', trip_counts=trip_counts)

    elif step == 'settings':
        return render_template(
            'admin/trips/builder/step_settings.html',
            title='Settings',
            trip=trip,
            current_step='settings',
            trip_counts=trip_counts,
        )

    abort(404)


@bp.route('/trips/<int:id>/delete', methods=['POST'])
@admin_required
def delete_trip(id):
    trip = Trip.query.get_or_404(id)
    # pending_bookings / messages 的 trip_id 为 NOT NULL；无 cascade 时 ORM 会尝试 SET NULL
    PendingBooking.query.filter_by(trip_id=trip.id).delete(synchronize_session=False)
    Message.query.filter_by(trip_id=trip.id).delete(synchronize_session=False)
    db.session.delete(trip)
    db.session.commit()
    flash('Trip deleted.')
    return redirect(url_for('admin.trips'))


@bp.route('/trips/<int:id>/copy', methods=['POST'])
@login_required
def copy_trip(id):
    import time

    trip = Trip.query.get_or_404(id)
    timestamp = int(time.time())
    new_status = copy_trip_status(trip)
    new_trip = Trip(
        title=f"Copy of {trip.title}",
        slug=f"{trip.slug}-copy-{timestamp}",
        price=trip.price,
        start_date=trip.start_date,
        end_date=trip.end_date,
        description=trip.description,
        hero_image=trip.hero_image,
        highlight_image=trip.highlight_image,
        capacity=trip.capacity,
        destination_text=trip.destination_text,
        trip_includes=trip.trip_includes,
        trip_excludes=trip.trip_excludes,
        status=new_status,
        is_published=False,
        color=trip.color,
    )
    db.session.add(new_trip)
    db.session.commit()
    # Draft copies stay Draft even if fields are complete (until next save/sync).
    target = 'draft' if new_status == 'draft' else 'unpublished'
    flash(
        'Trip copied as Draft.'
        if target == 'draft'
        else 'Trip copied as Unpublished (not live). Publish when ready.'
    )
    return redirect(url_for('admin.trips', filter=target))


@bp.route('/trips/<int:id>/publish', methods=['POST'])
@login_required
def publish_trip_route(id):
    trip = Trip.query.get_or_404(id)
    ok, msg = publish_trip(trip)
    db.session.commit()
    flash(msg, 'success' if ok else 'warning')
    return redirect(url_for('admin.trips', filter=trip_list_bucket(trip)))


@bp.route('/trips/<int:id>/unpublish', methods=['POST'])
@login_required
def unpublish_trip_route(id):
    trip = Trip.query.get_or_404(id)
    ok, msg = unpublish_trip(trip)
    db.session.commit()
    flash(msg, 'success' if ok else 'warning')
    return redirect(url_for('admin.trips', filter=trip_list_bucket(trip)))


# Back-compat aliases (old menus / bookmarks)
@bp.route('/trips/<int:id>/deactivate', methods=['POST'])
@login_required
def deactivate_trip(id):
    return unpublish_trip_route(id)


@bp.route('/trips/<int:id>/reactivate', methods=['POST'])
@login_required
def reactivate_trip(id):
    # Old "reactivate" meant go live — now only Publish from Unpublished
    trip = Trip.query.get_or_404(id)
    sync_trip_lifecycle(trip, commit=True)
    if trip.status == 'unpublished':
        return publish_trip_route(id)
    flash('Open the trip and use Publish when it is ready (Unpublished).', 'warning')
    return redirect(url_for('admin.trips', filter=trip_list_bucket(trip)))


@bp.route('/trips/<int:id>/archive', methods=['POST'])
@login_required
def archive_trip(id):
    # Legacy: treat archive as unpublish into unpublished/draft via lifecycle
    trip = Trip.query.get_or_404(id)
    if trip.status == 'published':
        ok, msg = unpublish_trip(trip)
        db.session.commit()
        flash(msg, 'success' if ok else 'warning')
    else:
        flash('Trip is not live.', 'warning')
    return redirect(url_for('admin.trips', filter=trip_list_bucket(trip)))


@bp.route('/cities')
@login_required
def cities():
    cities = City.query.all()
    return render_template('admin/cities/list.html', title='城市管理', cities=cities)


@bp.route('/cities/new', methods=['GET', 'POST'])
@login_required
def new_city():
    form = CityForm()
    if form.validate_on_submit():
        city = City(
            name=form.name.data,
            country=form.country.data,
            description=form.description.data,
            # image_url=form.image_url.data
        )
        db.session.add(city)
        db.session.commit()
        flash('城市已创建')
        return redirect(url_for('admin.cities'))
    return render_template('admin/cities/form.html', title='新建城市', form=form)


@bp.route('/cities/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_city(id):
    city = City.query.get_or_404(id)
    form = CityForm(obj=city)
    if request.method == 'GET':
        # form.image_url.data = city.image_url
        pass
    if form.validate_on_submit():
        city.name = form.name.data
        city.country = form.country.data
        city.description = form.description.data
        # city.image_url = form.image_url.data
        db.session.commit()
        flash('城市已更新')
        return redirect(url_for('admin.cities'))
    return render_template('admin/cities/form.html', title='编辑城市', form=form)


@bp.route('/cities/<int:id>/delete', methods=['POST'])
@login_required
def delete_city(id):
    city = City.query.get_or_404(id)
    db.session.delete(city)
    db.session.commit()
    flash('城市已删除')
    return redirect(url_for('admin.cities'))


ADMIN_LIST_PER_PAGE = 20


def _parse_lead_interest_list(lead):
    """解析 Lead.interest 为列表，挂到 lead.interest_list。"""
    if lead.interest:
        try:
            interest_data = json.loads(lead.interest)
            if isinstance(interest_data, list):
                lead.interest_list = interest_data
            else:
                lead.interest_list = [interest_data]
        except (json.JSONDecodeError, TypeError):
            if ',' in lead.interest:
                lead.interest_list = [item.strip() for item in lead.interest.split(',')]
            else:
                lead.interest_list = [lead.interest.strip()]
    else:
        lead.interest_list = []


def _apply_testimonial_source_filter(query, source_filter):
    """source=homepage 时把空 source 视为 homepage。"""
    if source_filter == 'homepage':
        return query.filter(
            or_(
                Testimonial.source == 'homepage',
                Testimonial.source.is_(None),
                Testimonial.source == '',
            )
        )
    if source_filter == 'feedback':
        return query.filter(Testimonial.source == 'feedback')
    return query


@bp.route('/customers')
@login_required
def customers():
    page = request.args.get('page', 1, type=int)
    pagination = (
        Client.query.order_by(Client.created_at.desc())
        .paginate(page=page, per_page=ADMIN_LIST_PER_PAGE, error_out=False)
    )
    clients = pagination.items

    # Calculate client stats (trips and collected amount) for current page only
    client_stats = {}
    for client in clients:
        bookings = client.bookings.all() if client.bookings else []

        total_collected = sum(b.amount_paid or 0.0 for b in bookings)

        trip_names = []
        seen_trip_ids = set()
        for booking in bookings:
            if booking.trip and booking.trip.id not in seen_trip_ids:
                trip_names.append(booking.trip.title)
                seen_trip_ids.add(booking.trip.id)

        client_stats[client.id] = {
            'trips': trip_names,
            'collected_amount': total_collected
        }

    return render_template(
        'admin/customers/customers.html',
        title='Customers',
        clients=clients,
        client_stats=client_stats,
        pagination=pagination,
    )


@bp.route('/customers/leads')
@login_required
def leads():
    page = request.args.get('page', 1, type=int)
    status_filter = request.args.get('status', 'all')
    search_q = (request.args.get('q') or '').strip()

    query = Lead.query
    if status_filter == 'new':
        query = query.filter(or_(Lead.status == 'new', Lead.status.is_(None), Lead.status == ''))
    elif status_filter in ('replied', 'converted', 'archived'):
        query = query.filter_by(status=status_filter)

    if search_q:
        like = f'%{search_q}%'
        query = query.filter(
            or_(
                Lead.name.ilike(like),
                Lead.email.ilike(like),
                Lead.phone.ilike(like),
                Lead.organization.ilike(like),
            )
        )

    pagination = query.order_by(Lead.created_at.desc()).paginate(
        page=page, per_page=ADMIN_LIST_PER_PAGE, error_out=False
    )
    leads_list = pagination.items
    for lead in leads_list:
        _parse_lead_interest_list(lead)

    stats = {
        'total': Lead.query.count(),
        'new': Lead.query.filter(or_(Lead.status == 'new', Lead.status.is_(None), Lead.status == '')).count(),
        'replied': Lead.query.filter_by(status='replied').count(),
        'converted': Lead.query.filter_by(status='converted').count(),
        'archived': Lead.query.filter_by(status='archived').count(),
    }

    return render_template(
        'admin/customers/leads.html',
        title='Leads',
        leads=leads_list,
        stats=stats,
        pagination=pagination,
        status_filter=status_filter,
        search_q=search_q,
    )


@bp.route('/customers/leads/<int:id>/update-status', methods=['POST'])
@login_required
def update_lead_status(id):
    """更新Lead状态"""
    lead = Lead.query.get_or_404(id)
    data = request.get_json()
    new_status = data.get('status')
    
    if new_status not in ['new', 'replied', 'converted', 'archived']:
        return jsonify({'success': False, 'message': '无效的状态值'}), 400
    
    lead.status = new_status
    db.session.commit()
    
    return jsonify({'success': True, 'message': '状态更新成功'})


@bp.route('/customers/leads/<int:id>/delete', methods=['POST'])
@admin_required
def delete_lead(id):
    """删除Lead"""
    lead = Lead.query.get_or_404(id)
    
    try:
        # Delete lead (no related data to cascade delete)
        db.session.delete(lead)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Lead deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting lead: {str(e)}")
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500


@bp.route('/customers/leads/bulk-delete', methods=['POST'])
@admin_required
def bulk_delete_leads():
    """批量删除 Lead（仅删除请求中的 id 列表）"""
    data = request.get_json(silent=True) or {}
    raw_ids = data.get('ids') or []
    try:
        ids = sorted({int(x) for x in raw_ids})
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Invalid ids'}), 400
    if not ids:
        return jsonify({'success': False, 'message': 'No leads selected'}), 400
    if len(ids) > 200:
        return jsonify({'success': False, 'message': 'Too many ids (max 200)'}), 400

    try:
        deleted = Lead.query.filter(Lead.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'success': True, 'message': f'Deleted {deleted} lead(s)', 'deleted': deleted})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error bulk deleting leads: {str(e)}")
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500


@bp.route('/customers/testimonials')
@login_required
def testimonials():
    """首页 Testimonials 管理"""
    status_filter = request.args.get('status', 'all')
    source_filter = request.args.get('source', 'all')
    page = request.args.get('page', 1, type=int)
    pagination = None

    def scoped(status=None):
        q = Testimonial.query
        if status:
            q = q.filter_by(status=status)
        return _apply_testimonial_source_filter(q, source_filter)

    if status_filter == 'approved':
        # 全量加载，保留拖拽排序
        items = (
            scoped('approved')
            .order_by(Testimonial.sort_order.asc(), Testimonial.id.asc())
            .all()
        )
    elif status_filter in ('pending', 'rejected'):
        pagination = (
            scoped(status_filter)
            .order_by(Testimonial.created_at.desc())
            .paginate(page=page, per_page=ADMIN_LIST_PER_PAGE, error_out=False)
        )
        items = pagination.items
    else:
        # All：approved 全量可拖；pending/rejected 分页
        approved = (
            scoped('approved')
            .order_by(Testimonial.sort_order.asc(), Testimonial.id.asc())
            .all()
        )
        inbox_q = (
            scoped()
            .filter(Testimonial.status.in_(['pending', 'rejected']))
            .order_by(
                case((Testimonial.status == 'pending', 0), else_=1),
                Testimonial.created_at.desc(),
            )
        )
        pagination = inbox_q.paginate(page=page, per_page=ADMIN_LIST_PER_PAGE, error_out=False)
        items = approved + pagination.items

    stats = {
        'total': Testimonial.query.count(),
        'pending': Testimonial.query.filter_by(status='pending').count(),
        'approved': Testimonial.query.filter_by(status='approved').count(),
        'rejected': Testimonial.query.filter_by(status='rejected').count(),
    }
    can_reorder = (
        status_filter in ('all', 'approved')
        and source_filter in ('all', 'homepage')
        and stats['approved'] > 0
    )

    return render_template(
        'admin/customers/testimonials.html',
        title='Testimonials',
        testimonials=items,
        stats=stats,
        status_filter=status_filter,
        source_filter=source_filter,
        can_reorder=can_reorder,
        pagination=pagination,
    )


def _validate_testimonial_payload(data, existing=None):
    quote = (data.get('quote') or '').strip()
    author_name = (data.get('author_name') or '').strip()
    organization = (data.get('organization') or '').strip() or None
    status = (data.get('status') or 'pending').strip()
    source = (existing.source if existing else data.get('source') or 'homepage').strip()
    max_quote = 2000 if source == 'feedback' else 500

    if status not in ('pending', 'approved', 'rejected'):
        return None, 'Invalid status.'
    if not author_name:
        return None, 'Author name is required.'
    if len(quote) < 20:
        return None, 'Quote must be at least 20 characters.'
    if len(quote) > max_quote:
        return None, f'Quote must be {max_quote} characters or fewer.'
    if len(author_name) > 128:
        return None, 'Author name is too long.'
    if organization and len(organization) > 200:
        return None, 'School or organization name is too long.'

    return {
        'quote': quote,
        'author_name': author_name,
        'organization': organization,
        'status': status,
        'source': source,
    }, None


@bp.route('/customers/testimonials/<int:id>/json')
@login_required
def testimonial_json(id):
    testimonial = Testimonial.query.get_or_404(id)
    return jsonify({
        'success': True,
        'testimonial': {
            'id': testimonial.id,
            'quote': testimonial.quote,
            'author_name': testimonial.author_name,
            'organization': testimonial.organization or '',
            'status': testimonial.status,
            'source': testimonial.source or 'homepage',
            'email': testimonial.email or '',
            'phone': testimonial.phone or '',
            'rating': testimonial.rating or '',
            'rating_label': testimonial.rating_label or '',
        },
    })


@bp.route('/customers/testimonials/save', methods=['POST'])
@login_required
def save_testimonial_api():
    data = request.get_json(silent=True) or {}
    testimonial_id = data.get('id')
    existing = Testimonial.query.get(int(testimonial_id)) if testimonial_id else None
    payload, error = _validate_testimonial_payload(data, existing=existing)
    if error:
        return jsonify({'success': False, 'message': error}), 400

    try:
        if existing:
            was_approved = existing.status == 'approved'
            existing.quote = payload['quote']
            existing.author_name = payload['author_name']
            existing.organization = payload['organization']
            existing.status = payload['status']
            if existing.status == 'approved' and not was_approved:
                assign_carousel_sort_order(existing)
            message = 'Testimonial updated.'
        else:
            testimonial = Testimonial(
                quote=payload['quote'],
                author_name=payload['author_name'],
                organization=payload['organization'],
                status=payload['status'],
                source='homepage',
            )
            db.session.add(testimonial)
            db.session.flush()
            if testimonial.status == 'approved':
                assign_carousel_sort_order(testimonial)
            message = 'Testimonial created.'
        db.session.commit()
        return jsonify({'success': True, 'message': message})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error saving testimonial: {str(e)}")
        return jsonify({'success': False, 'message': 'Unable to save testimonial.'}), 500


@bp.route('/customers/testimonials/new', methods=['GET', 'POST'])
@login_required
def new_testimonial():
    form = TestimonialForm()
    if form.validate_on_submit():
        testimonial = Testimonial(
            quote=form.quote.data.strip(),
            author_name=form.author_name.data.strip(),
            organization=(form.organization.data or '').strip() or None,
            status=form.status.data,
            source='homepage',
        )
        db.session.add(testimonial)
        db.session.flush()
        if testimonial.status == 'approved':
            assign_carousel_sort_order(testimonial)
        db.session.commit()
        flash('Testimonial created.')
        return redirect(url_for('admin.testimonials'))
    if request.method == 'GET':
        form.status.data = 'approved'
    return render_template('admin/customers/testimonial_form.html', title='New Testimonial', form=form)


@bp.route('/customers/testimonials/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_testimonial(id):
    testimonial = Testimonial.query.get_or_404(id)
    form = TestimonialForm(obj=testimonial)
    if form.validate_on_submit():
        was_approved = testimonial.status == 'approved'
        testimonial.quote = form.quote.data.strip()
        testimonial.author_name = form.author_name.data.strip()
        testimonial.organization = (form.organization.data or '').strip() or None
        testimonial.status = form.status.data
        if testimonial.status == 'approved' and not was_approved:
            assign_carousel_sort_order(testimonial)
        db.session.commit()
        flash('Testimonial updated.')
        return redirect(url_for('admin.testimonials'))
    return render_template('admin/customers/testimonial_form.html', title='Edit Testimonial', form=form, testimonial=testimonial)


@bp.route('/customers/testimonials/<int:id>/approve', methods=['POST'])
@login_required
def approve_testimonial(id):
    testimonial = Testimonial.query.get_or_404(id)
    testimonial.status = 'approved'
    assign_carousel_sort_order(testimonial)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Testimonial approved.'})


@bp.route('/customers/testimonials/reorder', methods=['POST'])
@login_required
def reorder_testimonials():
    """保存首页轮播拖拽顺序（仅 approved）。"""
    data = request.get_json() or {}
    ids = data.get('ids')
    if not isinstance(ids, list) or not ids:
        return jsonify({'success': False, 'message': 'Invalid order.'}), 400

    try:
        id_list = [int(item_id) for item_id in ids]
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Invalid order.'}), 400

    approved_ids = {
        row.id
        for row in Testimonial.query.filter_by(status='approved').all()
    }
    if set(id_list) != approved_ids:
        return jsonify({'success': False, 'message': 'Order must include all approved testimonials.'}), 400

    for index, testimonial_id in enumerate(id_list):
        testimonial = Testimonial.query.get(testimonial_id)
        if not testimonial:
            return jsonify({'success': False, 'message': 'Testimonial not found.'}), 404
        testimonial.sort_order = index

    db.session.commit()
    return jsonify({'success': True, 'message': 'Carousel order saved.'})


@bp.route('/customers/testimonials/<int:id>/reject', methods=['POST'])
@login_required
def reject_testimonial(id):
    testimonial = Testimonial.query.get_or_404(id)
    testimonial.status = 'rejected'
    db.session.commit()
    return jsonify({'success': True, 'message': 'Testimonial rejected.'})


@bp.route('/customers/testimonials/<int:id>/delete', methods=['POST'])
@login_required
def delete_testimonial(id):
    testimonial = Testimonial.query.get_or_404(id)
    try:
        db.session.delete(testimonial)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Testimonial deleted.'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting testimonial: {str(e)}")
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500


@bp.route('/customers/testimonials/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_testimonials():
    """批量删除 Testimonials（仅请求中的 id；含 approved 时需 confirm_approved）"""
    data = request.get_json(silent=True) or {}
    raw_ids = data.get('ids') or []
    confirm_approved = bool(data.get('confirm_approved'))
    try:
        ids = sorted({int(x) for x in raw_ids})
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Invalid ids'}), 400
    if not ids:
        return jsonify({'success': False, 'message': 'No testimonials selected'}), 400
    if len(ids) > 200:
        return jsonify({'success': False, 'message': 'Too many ids (max 200)'}), 400

    rows = Testimonial.query.filter(Testimonial.id.in_(ids)).all()
    if not rows:
        return jsonify({'success': False, 'message': 'No matching testimonials'}), 404
    approved_count = sum(1 for r in rows if r.status == 'approved')
    if approved_count and not confirm_approved:
        return jsonify({
            'success': False,
            'need_confirm_approved': True,
            'approved_count': approved_count,
            'message': f'{approved_count} selected item(s) are approved (homepage carousel). Confirm to delete.',
        }), 400

    try:
        deleted = Testimonial.query.filter(Testimonial.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({'success': True, 'message': f'Deleted {deleted} testimonial(s)', 'deleted': deleted})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error bulk deleting testimonials: {str(e)}")
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500


# Deprecated clients route, redirect to customers
@bp.route('/clients')
@login_required
def clients():
    return redirect(url_for('admin.customers'))


@bp.route('/clients/new', methods=['GET', 'POST'])
@login_required
def new_client():
    form = ClientForm()
    if form.validate_on_submit():
        client = Client(
            name=form.name.data,
            email=form.email.data,
            phone=form.phone.data,
            notes=form.notes.data
        )
        db.session.add(client)
        db.session.commit()
        flash('客户已创建')
        return redirect(url_for('admin.clients'))
    return render_template('admin/clients/form.html', title='新建客户', form=form)


@bp.route('/clients/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def edit_client(id):
    client = Client.query.get_or_404(id)
    form = ClientForm(obj=client)
    if form.validate_on_submit():
        client.name = form.name.data
        client.email = form.email.data
        client.phone = form.phone.data
        client.notes = form.notes.data
        db.session.commit()
        flash('客户已更新')
        return redirect(url_for('admin.clients'))
    return render_template('admin/clients/form.html', title='编辑客户', form=form)


@bp.route('/clients/<int:id>/delete', methods=['POST'])
@admin_required
def delete_client(id):
    client = Client.query.get_or_404(id)
    
    # Check if client has bookings
    bookings_count = client.bookings.count()
    if bookings_count > 0:
        # Delete all related bookings first (cascade delete)
        for booking in client.bookings:
            # Delete add-ons (从两个来源删除)
            for participant in booking.participants:
                for addon in participant.addons:
                    db.session.delete(addon)
                db.session.delete(participant)
            # Delete booking.addons (直接关联的 add-ons)
            for addon in booking.addons:
                db.session.delete(addon)
            # Delete booking packages
            for bp in booking.booking_packages:
                db.session.delete(bp)
            db.session.delete(booking)
    
    # Delete client
    db.session.delete(client)
    db.session.commit()
    flash('客户已删除')
    return redirect(url_for('admin.customers'))


@bp.route('/customers/<int:id>/delete', methods=['POST'])
@admin_required
def delete_customer(id):
    """删除客户（Customers 页面使用）"""
    client = Client.query.get_or_404(id)
    
    try:
        # Check if client has bookings
        bookings_count = client.bookings.count()
        if bookings_count > 0:
            # Delete all related bookings first (cascade delete)
            for booking in client.bookings:
                # Delete add-ons (从两个来源删除)
                for participant in booking.participants:
                    for addon in participant.addons:
                        db.session.delete(addon)
                    db.session.delete(participant)
                # Delete booking.addons (直接关联的 add-ons)
                for addon in booking.addons:
                    db.session.delete(addon)
                # Delete booking packages
                for bp in booking.booking_packages:
                    db.session.delete(bp)
                db.session.delete(booking)
        
        # Delete client
        db.session.delete(client)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Customer deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting customer: {str(e)}")
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500


@bp.route('/trips/<int:id>/checkout_test')
@admin_required
def checkout_test(id):
    trip = Trip.query.get_or_404(id)
    # 模拟一个成功和取消的 URL
    success_url = url_for('admin.trips', _external=True) + '?payment=success'
    cancel_url = url_for('admin.trips', _external=True) + '?payment=cancelled'
    
    # 使用当前管理员邮箱作为测试
    session = create_checkout_session(trip, 'admin@example.com', success_url, cancel_url)
    
    if session:
        return redirect(session.url)
    else:
        flash('无法创建支付会话，请检查 Stripe 配置')
        return redirect(url_for('admin.trips'))


@bp.route('/payments')
@login_required
def payments():
    """
    Payments：以 order 为基本单位。
    Full = 一次付全款 + 定金/单笔尾款（纯）+ 混单全款产品行；
    Installment = 定金后多期 + 混单分期 schedule（含混单的 deposit+final）。
    """
    from app.admin.payment_list_data import (
        build_mixed_full_portion_groups,
        build_one_time_order_groups,
        build_schedule_order_groups,
    )

    tab = request.args.get('tab', 'records')
    search = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '')

    if tab == 'records':
        # Full：纯定金+单笔尾款（排除混单 schedule）+ 一次付全款 + 混单全款产品行
        schedule_groups, today = build_schedule_order_groups(
            plan_kinds={'deposit_balance'},
            search=search,
            status_filter=status_filter,
            mixed_mode='exclude',
        )
        schedule_ids = {g['booking_id'] for g in schedule_groups}
        one_time_groups, _ = build_one_time_order_groups(
            search=search,
            status_filter=status_filter,
            exclude_booking_ids=schedule_ids,
        )
        mixed_full_groups, _ = build_mixed_full_portion_groups(
            search=search,
            status_filter=status_filter,
        )
        orders_grouped = schedule_groups + one_time_groups + mixed_full_groups
        orders_grouped.sort(
            key=lambda g: (
                (g.get('deposit') and g['deposit'].created_at)
                or (g.get('primary_payment') and g['primary_payment'].created_at)
                or datetime.min
            ),
            reverse=True,
        )
        return render_template(
            'admin/payments/list.html',
            title='Payments',
            tab=tab,
            orders_grouped=orders_grouped,
            today=today,
            search=search,
            status_filter=status_filter,
        )

    if tab == 'installments':
        # 多期 + 混单的定金/单笔尾款（纯 deposit_balance 仍在 Full）
        multi_groups, today = build_schedule_order_groups(
            plan_kinds={'multi'},
            search=search,
            status_filter=status_filter,
            mixed_mode='include',
        )
        mixed_balance_groups, _ = build_schedule_order_groups(
            plan_kinds={'deposit_balance'},
            search=search,
            status_filter=status_filter,
            mixed_mode='only',
        )
        orders_grouped = multi_groups + mixed_balance_groups
        orders_grouped.sort(
            key=lambda g: (
                (g.get('deposit') and g['deposit'].created_at)
                or datetime.min
            ),
            reverse=True,
        )
        return render_template(
            'admin/payments/list.html',
            title='Payments',
            tab=tab,
            orders_grouped=orders_grouped,
            today=today,
            search=search,
            status_filter=status_filter,
        )

    return redirect(url_for('admin.payments', tab='records'))


@bp.route('/reports')
@login_required
def reports():
    """
    Reports页面 - Financial Overview
    """
    from datetime import date

    today = date.today()
    search = request.args.get('search', '').strip()

    trips_query = Trip.query.filter(Trip.status != 'draft')
    if search:
        trips_query = trips_query.filter(Trip.title.ilike(f'%{search}%'))

    all_trips = trips_query.all()

    upcoming_trips = [t for t in all_trips if t.status == 'published' and t.end_date and t.end_date >= today]
    past_trips = [t for t in all_trips if t.status == 'published' and t.end_date and t.end_date < today]
    deactivated_trips = [t for t in all_trips if t.status in ('unpublished', 'deactivated')]

    def get_trip_summary(trip):
        stats = calculate_trip_stats(trip)
        bookings = trip.bookings.all() if trip.bookings else []
        return {
            'trip': trip,
            'booking_count': len(bookings),
            'participants_count': stats['participants_count'],
            'amount_gross': stats['amount_gross'],
            'amount_discount': stats['amount_discount'],
            'amount_expected': stats['amount_expected'],
            'amount_paid': stats['amount_paid'],
            'amount_refunded': stats.get('amount_refunded', 0.0) or 0.0,
            'amount_pending': stats['amount_available']
        }

    upcoming_summaries = sorted([get_trip_summary(t) for t in upcoming_trips],
                                key=lambda x: x['trip'].start_date or date.min, reverse=True)
    past_summaries = sorted([get_trip_summary(t) for t in past_trips],
                            key=lambda x: x['trip'].end_date or date.min, reverse=True)
    deactivated_summaries = sorted([get_trip_summary(t) for t in deactivated_trips],
                                   key=lambda x: x['trip'].updated_at or datetime.min, reverse=True)

    def calculate_category_total(summaries):
        return {
            'total_trips': len(summaries),
            'total_bookings': sum(s['booking_count'] for s in summaries),
            'total_participants': sum(s['participants_count'] for s in summaries),
            'total_gross': sum(s['amount_gross'] for s in summaries),
            'total_discount': sum(s['amount_discount'] for s in summaries),
            'total_expected': sum(s['amount_expected'] for s in summaries),
            'total_paid': sum(s['amount_paid'] for s in summaries),
            'total_refunded': sum(s.get('amount_refunded', 0.0) or 0.0 for s in summaries),
            'total_pending': sum(s['amount_pending'] for s in summaries)
        }

    upcoming_total = calculate_category_total(upcoming_summaries)
    past_total = calculate_category_total(past_summaries)
    deactivated_total = calculate_category_total(deactivated_summaries)

    grand_total = {
        'total_trips': len(all_trips),
        'total_bookings': upcoming_total['total_bookings'] + past_total['total_bookings'] + deactivated_total['total_bookings'],
        'total_participants': upcoming_total['total_participants'] + past_total['total_participants'] + deactivated_total['total_participants'],
        'total_gross': upcoming_total['total_gross'] + past_total['total_gross'] + deactivated_total['total_gross'],
        'total_discount': upcoming_total['total_discount'] + past_total['total_discount'] + deactivated_total['total_discount'],
        'total_expected': upcoming_total['total_expected'] + past_total['total_expected'] + deactivated_total['total_expected'],
        'total_paid': upcoming_total['total_paid'] + past_total['total_paid'] + deactivated_total['total_paid'],
        'total_refunded': upcoming_total['total_refunded'] + past_total['total_refunded'] + deactivated_total['total_refunded'],
        'total_pending': upcoming_total['total_pending'] + past_total['total_pending'] + deactivated_total['total_pending']
    }

    return render_template(
        'admin/reports/financial_overview.html',
        title='Reports',
        search=search,
        grand_total=grand_total,
        upcoming_summaries=upcoming_summaries,
        upcoming_total=upcoming_total,
        past_summaries=past_summaries,
        past_total=past_total,
        deactivated_summaries=deactivated_summaries,
        deactivated_total=deactivated_total
    )


@bp.route('/payments/api')
@login_required
def payments_api():
    """Payments API端点，返回JSON格式数据用于自动更新（重构版：支持 Booking 关联）"""
    # 预加载关联数据
    query = Payment.query.options(
        joinedload(Payment.booking),
        joinedload(Payment.client),
        joinedload(Payment.trip)
    )
    
    # 搜索功能
    search = request.args.get('search', '').strip()
    if search:
        query = query.join(Booking, isouter=True).join(Client, isouter=True).join(Trip, isouter=True).filter(
            or_(
                Booking.id.cast(db.String).ilike(f'%{search}%'),
                Client.name.ilike(f'%{search}%'),
                Client.email.ilike(f'%{search}%'),
                Trip.title.ilike(f'%{search}%')
            )
        )
    
    # 状态筛选
    status_filter = request.args.get('status', '')
    if status_filter:
        query = query.filter(Payment.status == status_filter)
    
    # 排序
    payments = query.order_by(Payment.created_at.desc()).limit(100).all()
    
    # 转换为JSON格式
    payments_data = []
    for payment in payments:
        payments_data.append({
            'id': payment.id,
            'booking_id': payment.booking_id,
            'client_name': payment.client.name if payment.client else (payment.booking.buyer_name if payment.booking else None),
            'client_email': payment.client.email if payment.client else (payment.booking.buyer_email if payment.booking else None),
            'trip_title': payment.trip.title if payment.trip else None,
            'amount': float(payment.amount) if payment.amount else 0.0,
            'status': payment.status or 'pending',
            'stripe_payment_intent_id': payment.stripe_payment_intent_id,
            'stripe_checkout_session_id': payment.stripe_checkout_session_id,
            'date': payment.created_at.strftime('%Y-%m-%d %H:%M') if payment.created_at else None,
            'paid_at': payment.paid_at.strftime('%Y-%m-%d %H:%M') if payment.paid_at else None
        })
    
    return jsonify({
        'success': True,
        'payments': payments_data,
        'count': len(payments_data)
    })


@bp.route('/payments/<int:payment_id>')
@login_required
def payment_detail(payment_id):
    """支付详情页面"""
    payment = Payment.query.options(
        joinedload(Payment.booking).joinedload(Booking.discount_code),
        joinedload(Payment.client),
        joinedload(Payment.trip),
        joinedload(Payment.installment_payment)
    ).get_or_404(payment_id)

    return render_template('admin/payments/detail.html',
                         title=f'Payment #{payment_id}',
                         payment=payment)


@bp.route('/payments/installments/api')
@login_required
def installment_payments_api():
    """分期付款列表 API"""
    query = InstallmentPayment.query.options(
        joinedload(InstallmentPayment.booking).joinedload(Booking.trip),
        joinedload(InstallmentPayment.booking).joinedload(Booking.client)
    )
    
    # 状态筛选
    status_filter = request.args.get('status', '')
    if status_filter:
        query = query.filter(InstallmentPayment.status == status_filter)
    
    # Booking ID 筛选
    booking_id = request.args.get('booking_id')
    if booking_id:
        query = query.filter(InstallmentPayment.booking_id == booking_id)
    
    # 按到期日期排序
    installments = query.order_by(InstallmentPayment.due_date.asc()).all()
    
    installments_data = []
    for inst in installments:
        installments_data.append({
            'id': inst.id,
            'booking_id': inst.booking_id,
            'booking_number': (inst.booking.order_number if inst.booking and inst.booking.order_number else (inst.booking.id if inst.booking else None)),
            'order_number': inst.booking.order_number if inst.booking else None,
            'trip_title': inst.booking.trip.title if inst.booking and inst.booking.trip else None,
            'buyer_name': inst.booking.buyer_name if inst.booking else None,
            'buyer_email': inst.booking.buyer_email if inst.booking else None,
            'installment_number': inst.installment_number,
            'amount': float(inst.amount) if inst.amount else 0.0,
            'due_date': inst.due_date.strftime('%Y-%m-%d') if inst.due_date else None,
            'status': inst.status or 'pending',
            'reminder_sent': inst.reminder_sent,
            'reminder_count': inst.reminder_count or 0,
            'paid_at': inst.paid_at.strftime('%Y-%m-%d %H:%M') if inst.paid_at else None
        })
    
    return jsonify({
        'success': True,
        'installments': installments_data,
        'count': len(installments_data)
    })


@bp.route('/payments/installments/<int:installment_id>/payment-link', methods=['GET'])
@admin_required
def installment_payment_link(installment_id):
    """Generate signed customer payment + payoff URLs for an unpaid installment."""
    installment = InstallmentPayment.query.get_or_404(installment_id)
    from app.installment_admin_links import (
        build_installment_payment_urls,
        installment_unpaid_action_error,
    )

    err = installment_unpaid_action_error(installment)
    if err:
        return jsonify({'success': False, 'error': err}), 400

    payment_url, payoff_url = build_installment_payment_urls(installment)
    return jsonify({
        'success': True,
        'payment_url': payment_url,
        'payoff_url': payoff_url,
        'installment_id': installment.id,
        'order_number': getattr(installment.booking, 'order_number', None),
    })


@bp.route('/bookings/<int:booking_id>/addons/catalog', methods=['GET'])
@admin_required
def booking_addons_catalog(booking_id):
    """Trip add-ons + participants for the Add add-on modal."""
    booking = Booking.query.get_or_404(booking_id)
    trip_addons = []
    if booking.trip_id:
        for ta in TripAddOn.query.filter_by(trip_id=booking.trip_id).order_by(TripAddOn.name).all():
            trip_addons.append({
                'id': ta.id,
                'name': ta.name,
                'price': float(ta.price or 0),
            })
    participants = []
    for p in booking.participants:
        if (getattr(p, 'status', None) or 'active') == 'withdrawn':
            continue
        participants.append({'id': p.id, 'name': p.name or f'Participant #{p.id}'})
    return jsonify({
        'success': True,
        'trip_addons': trip_addons,
        'participants': participants,
        'order_number': booking.order_number or f'#{booking.id}',
        'cancelled': booking.status == 'cancelled',
    })


@bp.route('/bookings/<int:booking_id>/addons', methods=['POST'])
@admin_required
def admin_add_booking_addon(booking_id):
    """Admin post-add an add-on (starts unpaid) and email the customer a pay link."""
    booking = Booking.query.get_or_404(booking_id)
    data = request.get_json(silent=True) or {}
    from app.addon_admin import (
        create_manual_booking_addon,
        send_addon_payment_email,
        serialize_booking_addon,
    )

    ba, err = create_manual_booking_addon(
        booking,
        trip_addon_id=data.get('addon_id'),
        quantity=data.get('quantity', 1),
        participant_id=data.get('participant_id'),
    )
    if err:
        return jsonify({'success': False, 'error': err}), 400
    db.session.commit()

    email_ok, email_msg = send_addon_payment_email(ba)
    if email_ok:
        message = f'Add-on added. Payment email sent ({email_msg}).'
    else:
        message = (
            f'Add-on added, but payment email failed: {email_msg or "unknown error"}. '
            'Use Actions → Send payment email to retry.'
        )
    return jsonify({
        'success': True,
        'addon': serialize_booking_addon(ba),
        'email_sent': bool(email_ok),
        'message': message,
    })


@bp.route('/booking-addons/<int:booking_addon_id>/payment-link', methods=['GET'])
@admin_required
def booking_addon_payment_link(booking_addon_id):
    """Signed customer pay URL for a manual unpaid add-on."""
    ba = BookingAddOn.query.get_or_404(booking_addon_id)
    from app.addon_admin import addon_payment_url
    from app.payments import booking_has_processing_ach_payment

    source = (getattr(ba, 'source', None) or 'booking')
    status = (getattr(ba, 'payment_status', None) or 'paid')
    if source != 'admin_manual':
        return jsonify({'success': False, 'error': 'Only manually added add-ons use this link'}), 400
    if status == 'paid':
        return jsonify({'success': False, 'error': 'This add-on is already paid'}), 400
    if booking_has_processing_ach_payment(ba.booking_id):
        return jsonify({
            'success': False,
            'error': 'A bank transfer for this order is still processing',
        }), 400
    return jsonify({
        'success': True,
        'payment_url': addon_payment_url(ba),
        'booking_addon_id': ba.id,
        'order_number': getattr(ba.booking, 'order_number', None),
    })


@bp.route('/booking-addons/<int:booking_addon_id>/payment-link/send', methods=['POST'])
@admin_required
def booking_addon_payment_link_send(booking_addon_id):
    """Email customer the add-on payment link."""
    ba = BookingAddOn.query.get_or_404(booking_addon_id)
    from app.addon_admin import send_addon_payment_email
    from app.payments import booking_has_processing_ach_payment

    source = (getattr(ba, 'source', None) or 'booking')
    status = (getattr(ba, 'payment_status', None) or 'paid')
    if source != 'admin_manual':
        return jsonify({'success': False, 'error': 'Only manually added add-ons can be emailed'}), 400
    if status == 'paid':
        return jsonify({'success': False, 'error': 'This add-on is already paid'}), 400
    if booking_has_processing_ach_payment(ba.booking_id):
        return jsonify({
            'success': False,
            'error': 'A bank transfer for this order is still processing',
        }), 400
    ok, msg = send_addon_payment_email(ba)
    if not ok:
        return jsonify({'success': False, 'error': msg or 'Failed to send'}), 400
    return jsonify({'success': True, 'message': msg})


@bp.route('/bookings/<int:booking_id>/auto-pay-link', methods=['GET'])
@admin_required
def admin_auto_pay_link(booking_id):
    """Return tokenized customer Auto Pay manage URL."""
    booking = Booking.query.get_or_404(booking_id)
    from app.auto_pay import auto_pay_manage_url, booking_has_installment_plan

    if not booking_has_installment_plan(booking):
        return jsonify({'success': False, 'error': 'No installment plan on this booking'}), 400
    return jsonify({
        'success': True,
        'auto_pay_url': auto_pay_manage_url(booking),
        'auto_pay_enabled': bool(booking.auto_pay_enabled),
    })


@bp.route('/bookings/<int:booking_id>/auto-pay-methods', methods=['GET'])
@admin_required
def admin_auto_pay_methods(booking_id):
    """List saved payment methods for Turn on picker + Auto Pay URL / buyer email."""
    booking = Booking.query.get_or_404(booking_id)
    from app.auto_pay import (
        auto_pay_manage_url,
        booking_has_installment_plan,
        payment_methods_for_booking,
    )

    if not booking_has_installment_plan(booking):
        return jsonify({'success': False, 'error': 'No installment plan on this booking'}), 400

    buyer_email = (booking.buyer_email or '').strip()
    if not buyer_email and booking.client:
        buyer_email = (booking.client.email or '').strip()

    return jsonify({
        'success': True,
        'payment_methods': payment_methods_for_booking(booking),
        'default_payment_method_id': getattr(booking, 'auto_pay_payment_method_id', None),
        'auto_pay_url': auto_pay_manage_url(booking),
        'auto_pay_enabled': bool(booking.auto_pay_enabled),
        'buyer_email': buyer_email or None,
        'order_number': booking.order_number or f'#{booking.id}',
    })


@bp.route('/bookings/<int:booking_id>/auto-pay-link/send', methods=['POST'])
@admin_required
def admin_send_auto_pay_link(booking_id):
    """Email the customer their Auto Pay enable/manage link."""
    booking = Booking.query.get_or_404(booking_id)
    from app.auto_pay import booking_has_installment_plan, send_auto_pay_invite_email

    if not booking_has_installment_plan(booking):
        return jsonify({'success': False, 'error': 'No installment plan on this booking'}), 400

    ok, msg = send_auto_pay_invite_email(booking)
    if not ok:
        return jsonify({'success': False, 'error': msg or 'Failed to send'}), 400
    return jsonify({'success': True, 'message': msg})


@bp.route('/bookings/<int:booking_id>/auto-pay', methods=['POST'])
@admin_required
def admin_toggle_auto_pay(booking_id):
    """Admin enable/disable Auto Pay. Enable requires a default PM (or sends Enable link)."""
    booking = Booking.query.get_or_404(booking_id)
    data = request.get_json(silent=True) or {}
    enable = data.get('enabled')
    if enable is None:
        return jsonify({'success': False, 'error': 'enabled true/false required'}), 400

    from app.auto_pay import (
        auto_pay_manage_url,
        booking_has_installment_plan,
        disable_auto_pay,
        enable_auto_pay,
    )

    if not booking_has_installment_plan(booking):
        return jsonify({'success': False, 'error': 'No installment plan on this booking'}), 400

    if enable:
        pm_id = (data.get('payment_method_id') or booking.auto_pay_payment_method_id or '').strip()
        if not pm_id:
            link = auto_pay_manage_url(booking)
            return jsonify({
                'success': False,
                'error': 'Select a saved payment method, or send the customer an Auto Pay link.',
                'auto_pay_url': link,
                'needs_customer_setup': True,
            }), 400
        ok, err = enable_auto_pay(booking, pm_id, source='admin')
        if not ok:
            return jsonify({'success': False, 'error': err or 'Could not enable'}), 400
        db.session.commit()
        return jsonify({
            'success': True,
            'auto_pay_enabled': True,
            'message': 'Auto Pay enabled',
            'auto_pay_url': auto_pay_manage_url(booking),
        })

    ok, err = disable_auto_pay(booking, source='admin')
    if not ok:
        return jsonify({'success': False, 'error': err or 'Could not disable'}), 400
    db.session.commit()
    return jsonify({
        'success': True,
        'auto_pay_enabled': False,
        'message': 'Auto Pay disabled',
        'auto_pay_url': auto_pay_manage_url(booking),
    })


@bp.route('/payments/installments/<int:installment_id>/send-reminder', methods=['POST'])
@admin_required
def send_installment_reminder(installment_id):
    """发送分期付款提醒邮件（已付清的期数不可催缴）"""
    installment = InstallmentPayment.query.get_or_404(installment_id)
    from app.installment_admin_links import installment_unpaid_action_error

    err = installment_unpaid_action_error(installment)
    if err:
        return jsonify({'success': False, 'error': err}), 400

    try:
        from app.tasks import send_installment_reminder_email, send_overdue_reminder_email
        from app.utils import pacific_today

        today = pacific_today()
        due = installment.due_date
        if due and due < today:
            ok = send_overdue_reminder_email(installment, (today - due).days)
        elif due:
            ok = send_installment_reminder_email(
                installment, days_until_due=max(0, (due - today).days)
            )
        else:
            ok = send_installment_reminder_email(installment, days_until_due=3)

        if not ok:
            return jsonify({'success': False, 'error': 'Failed to send reminder email'}), 500

        installment.reminder_sent = True
        installment.reminder_sent_at = datetime.utcnow()
        installment.reminder_count = (installment.reminder_count or 0) + 1
        if due and due < today and installment.status == 'pending':
            installment.status = 'overdue'
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Reminder email sent successfully'
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Failed to send reminder: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Failed to send reminder: {str(e)}'
        }), 500


@bp.route('/payments/installments/<int:installment_id>/mark-paid', methods=['POST'])
@admin_required
def mark_installment_paid(installment_id):
    """手动标记分期付款为已支付（线下）：写 Payment 账本，并禁止跳期。"""
    installment = InstallmentPayment.query.get_or_404(installment_id)
    
    if installment.status == 'paid':
        return jsonify({
            'success': False,
            'error': 'Installment already marked as paid'
        }), 400
    if (installment.status or '') == 'cancelled':
        return jsonify({
            'success': False,
            'error': 'This installment is cancelled',
        }), 400
    if Payment.query.filter_by(
        installment_payment_id=installment.id,
        status='succeeded',
    ).first():
        return jsonify({
            'success': False,
            'error': 'This installment already has a successful payment',
        }), 400

    # 与客户侧 catch-up 一致：须先付清更早未付期
    earlier_unpaid = (
        InstallmentPayment.query.filter(
            InstallmentPayment.booking_id == installment.booking_id,
            InstallmentPayment.status.in_(('pending', 'overdue')),
            InstallmentPayment.installment_number < (installment.installment_number or 0),
        )
        .order_by(InstallmentPayment.installment_number.asc())
        .first()
    )
    if earlier_unpaid:
        return jsonify({
            'success': False,
            'error': (
                f'Mark earlier installment #{earlier_unpaid.installment_number} '
                'paid first (or use catch-up payment link).'
            ),
        }), 400
    
    try:
        from app.payments import (
            calculate_booking_total,
            cancel_unpaid_installments,
            safe_cancel_payment_intent,
            void_stale_pending_payments,
        )
        
        if getattr(installment, 'payment_intent_id', None):
            safe_cancel_payment_intent(
                installment.payment_intent_id,
                reason=f'admin mark-paid installment {installment.id}',
            )

        paid_at = datetime.utcnow()
        installment.status = 'paid'
        installment.paid_at = paid_at
        
        booking = installment.booking
        amount = round(float(installment.amount or 0.0), 2)
        cents = int(round(amount * 100))
        step = 'initial' if (installment.installment_number or 0) == 0 else 'installment'
        payment = Payment(
            booking_id=booking.id,
            client_id=booking.client_id,
            trip_id=booking.trip_id,
            amount=amount,
            currency='usd',
            status='succeeded',
            installment_payment_id=installment.id,
            paid_at=paid_at,
            base_amount_cents=cents,
            fee_cents=0,
            tax_amount_cents=0,
            final_amount_cents=cents,
            payment_method_type='manual',
            payment_metadata={
                'payment_step': step,
                'payment_plan': 'installment',
                'source': 'admin_mark_paid',
                'manual': True,
                'installment_id': installment.id,
                'installment_number': installment.installment_number,
            },
        )
        db.session.add(payment)

        booking.amount_paid = round((booking.amount_paid or 0.0) + amount, 2)
        
        total_info = calculate_booking_total(booking)
        if booking.amount_paid + 0.001 >= float(total_info.get('total') or 0.0):
            booking.status = 'fully_paid'
            for bp in booking.booking_packages:
                bp.status = 'fully_paid'
            cancel_unpaid_installments(booking)
            void_stale_pending_payments(booking)
        elif booking.status == 'pending':
            booking.status = 'deposit_paid'
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Installment marked as paid',
            'payment_id': payment.id,
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Failed to mark installment as paid: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Failed to mark as paid: {str(e)}'
        }), 500


@bp.route('/payments/export')
@admin_required
def export_payments():
    """导出 Payments 为 Excel（列内容独立；视觉与 Manage Download Excel 统一）。"""
    try:
        from io import BytesIO
        from datetime import datetime
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = 'Payments'

        # 与 Manage bookings export 同款样式
        header_fill = PatternFill(start_color='3D3D3D', end_color='3D3D3D', fill_type='solid')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        body_font = Font(name='Arial', size=11)
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_align = Alignment(horizontal='left', vertical='center', wrap_text=False)
        money_align = Alignment(horizontal='right', vertical='center', wrap_text=False)
        thin = Border(
            left=Side(style='thin', color='E5E7EB'),
            right=Side(style='thin', color='E5E7EB'),
            top=Side(style='thin', color='E5E7EB'),
            bottom=Side(style='thin', color='E5E7EB'),
        )
        money_format = '#,##0.00'
        # Base / Fee / Charged / Net，以及退款块：Refunded / Reason / Refunded At
        money_cols = {5, 6, 7, 11, 12}
        dash = '-'

        # 付款信息在前（含 Net）；退款相关列集中在末尾
        headers = [
            'Order Number',
            'Client Name',
            'Client Email',
            'Trip Title',
            'Base Amount',
            'Card Fee',
            'Charged',
            'Status',
            'Paid At',
            'Created',
            'Net (base)',
            'Refunded (base)',
            'Refund Reason',
            'Refunded At',
        ]
        ws.row_dimensions[1].height = 28
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin

        from app.payments import (
            payment_base_amount,
            payment_charged_amount,
            payment_fee_amount,
            payment_refunded_clamped,
        )

        # 按付款时间升序（无 paid_at 则用 created_at），与 Payment History 一致
        payments = Payment.query.options(
            joinedload(Payment.client),
            joinedload(Payment.trip),
            joinedload(Payment.booking),
        ).order_by(
            func.coalesce(Payment.paid_at, Payment.created_at).asc(),
            Payment.id.asc(),
        ).all()

        for row_idx, payment in enumerate(payments, start=2):
            booking = payment.booking
            order_number = (
                (booking.order_number if booking else None)
                or (f'#{payment.booking_id}' if payment.booking_id else dash)
            )
            base = payment_base_amount(payment)
            fee = payment_fee_amount(payment)
            charged = payment_charged_amount(payment)
            refunded = payment_refunded_clamped(payment)
            net = round(max(0.0, base - refunded), 2)
            reason = (payment.refund_reason or '').strip()
            if not reason:
                reason = dash
            row_data = [
                order_number,
                payment.client.name if payment.client else dash,
                payment.client.email if payment.client else dash,
                payment.trip.title if payment.trip else dash,
                base,
                fee,
                charged,
                (payment.status or 'pending').replace('_', ' '),
                payment.paid_at.strftime('%Y-%m-%d %H:%M') if payment.paid_at else dash,
                payment.created_at.strftime('%Y-%m-%d %H:%M') if payment.created_at else dash,
                net,
                refunded if refunded else dash,
                reason,
                payment.refunded_at.strftime('%Y-%m-%d %H:%M') if payment.refunded_at else dash,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = body_font
                cell.border = thin
                if col_idx in money_cols:
                    cell.alignment = money_align
                    if isinstance(value, (int, float)):
                        cell.number_format = money_format
                else:
                    cell.alignment = cell_align

        def _display_width(value):
            if value is None:
                return 0
            width = 0
            for ch in str(value):
                width += 2 if ord(ch) > 0x2E80 else 1
            return width

        for col_idx in range(1, (ws.max_column or 0) + 1):
            max_len = 0
            for r in range(1, (ws.max_row or 0) + 1):
                max_len = max(max_len, _display_width(ws.cell(row=r, column=col_idx).value))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 3, 8), 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'payments_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        flash(f'导出失败: {str(e)}', 'error')
        return redirect(url_for('admin.payments'))




@bp.route('/trips/<int:id>/bookings/export')
@admin_required
def export_bookings(id):
    """导出行程预订快照 Excel（静态预填；含退款与取消信息，风格参考 WeTravel）。"""
    import io
    from flask import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import datetime

    trip = Trip.query.get_or_404(id)
    bookings = list(trip.bookings.order_by(Booking.created_at.asc()).all())
    from app.payments import booking_payment_display_status, booking_payment_type_display

    wb = Workbook()
    wb.remove(wb.active)

    # WeTravel 风格：深灰表头 + 白字；正文 Arial
    header_fill = PatternFill(start_color='3D3D3D', end_color='3D3D3D', fill_type='solid')
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    body_font = Font(name='Arial', size=11)
    body_bold = Font(name='Arial', size=11, bold=True)
    cancelled_font = Font(name='Arial', size=11, color='9CA3AF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(horizontal='left', vertical='center', wrap_text=False)
    money_align = Alignment(horizontal='right', vertical='center', wrap_text=False)
    thin = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB'),
    )
    money_format = '#,##0.00'
    dash = '-'

    def _display_width(value):
        if value is None:
            return 0
        width = 0
        for ch in str(value):
            o = ord(ch)
            width += 2 if o > 0x2E80 else 1
        return width

    def _autosize(ws, min_width=8, max_width=120):
        from openpyxl.utils import get_column_letter
        for col_idx in range(1, (ws.max_column or 0) + 1):
            max_len = 0
            for row_idx in range(1, (ws.max_row or 0) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                max_len = max(max_len, _display_width(cell.value))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max(max_len + 3, min_width), max_width
            )

    def _write_header(ws, headers):
        ws.row_dimensions[1].height = 28
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin

    def _booking_expected(booking):
        booking_gross = 0.0
        has_packages = False
        seen_addon_ids = set()
        for bp in booking.booking_packages:
            if bp.package:
                package_price = booking_package_unit_price(bp)
                quantity = int(bp.quantity) if bp.quantity is not None else 1
                booking_gross += package_price * quantity
                has_packages = True
        for participant in booking.participants:
            for booking_addon in participant.addons:
                if booking_addon.addon and booking_addon.id not in seen_addon_ids:
                    addon_price = booking_addon_unit_price(booking_addon)
                    quantity = int(booking_addon.quantity) if booking_addon.quantity is not None else 0
                    booking_gross += addon_price * quantity
                    seen_addon_ids.add(booking_addon.id)
        for booking_addon in booking.addons:
            if booking_addon.addon and booking_addon.id not in seen_addon_ids:
                addon_price = booking_addon_unit_price(booking_addon)
                quantity = int(booking_addon.quantity) if booking_addon.quantity is not None else 0
                booking_gross += addon_price * quantity
                seen_addon_ids.add(booking_addon.id)
        discount = float(booking.discount_amount) if booking.discount_amount else 0.0
        if not has_packages:
            return round(float(booking.amount_paid) if booking.amount_paid is not None else 0.0, 2)
        return round(max(0.0, booking_gross - discount), 2)

    def _booking_payment_totals(booking):
        """返回 (payments_received 实扣合计, card_fees 卡费合计, refunds 已退合计)。"""
        from app.payments import payment_fee_amount

        received = 0.0
        card_fees = 0.0
        refunds = 0.0
        for payment in Payment.query.filter_by(booking_id=booking.id).all():
            if payment.status in ('succeeded', 'partially_refunded', 'refunded'):
                received += float(payment.amount or 0.0)
                card_fees += payment_fee_amount(payment)
                refunds += float(payment.refunded_amount or 0.0)
        return round(received, 2), round(card_fees, 2), round(refunds, 2)

    def _participant_names(booking, mark_cancelled=False):
        names = []
        for p in booking.participants:
            n = (p.name or '').strip() or (
                f"{getattr(p, 'first_name', '') or ''} {getattr(p, 'last_name', '') or ''}".strip()
            )
            if not n:
                continue
            if (getattr(p, 'status', None) or 'active') == 'withdrawn':
                names.append(f'{n} (withdrawn)')
            elif mark_cancelled or booking.status == 'cancelled':
                names.append(f'{n} (cancelled)')
            else:
                names.append(n)
        return '; '.join(names) if names else dash

    # ===== Sheet 1: Participants（活跃订单）=====
    ws_participants = wb.create_sheet("Participants")
    participants_headers, participants_rows = _get_participants_export_data(
        trip, include_cancelled=False
    )
    _write_header(ws_participants, participants_headers)
    for row_idx, row_data in enumerate(participants_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_participants.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = cell_align
            cell.border = thin
    _autosize(ws_participants)

    # ===== Sheet 2: Contact =====
    ws_contact = wb.create_sheet("Contact")
    contact_headers = [
        'Order number', 'Sign up #', 'Contact First Name', 'Contact Last Name',
        'Attendee', 'Contact Email', 'Phone', 'Address', 'City', 'State',
        'ZIP', 'Home Phone', 'Emergency Contact Name', 'EC Email', 'EC Phone', 'Status',
    ]
    _write_header(ws_contact, contact_headers)

    contact_row = 2
    active_idx = 0
    for booking in bookings:
        if booking.status == 'cancelled':
            continue
        active_idx += 1
        client = booking.client
        buyer_first = booking.buyer_first_name or ''
        buyer_last = booking.buyer_last_name or ''
        if not buyer_first and not buyer_last and client:
            parts = (client.name or '').strip().split(None, 1)
            buyer_first = parts[0] if parts else ''
            buyer_last = parts[1] if len(parts) > 1 else ''
        attendees_str = _participant_names(booking)
        order_no = booking.order_number or f'#{booking.id}'
        contact_row_data = [
            order_no,
            active_idx,
            buyer_first or dash,
            buyer_last or dash,
            attendees_str,
            booking.buyer_email or (client.email if client else None) or dash,
            booking.buyer_phone or (client.phone if client else None) or dash,
            booking.buyer_address or dash,
            booking.buyer_city or dash,
            booking.buyer_state or dash,
            booking.buyer_zip_code or dash,
            booking.buyer_home_phone or booking.buyer_phone or (client.phone if client else None) or dash,
            booking.buyer_emergency_contact_name or dash,
            booking.buyer_emergency_contact_email or dash,
            booking.buyer_emergency_contact_phone or dash,
            (booking_payment_display_status(booking) or booking.status or '').replace('_', ' ').title(),
        ]
        for col_idx, value in enumerate(contact_row_data, start=1):
            cell = ws_contact.cell(row=contact_row, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = cell_align
            cell.border = thin
        contact_row += 1
    _autosize(ws_contact)

    # ===== Sheet 3: Bookings Summary（含 Refunds，取消单也列出）=====
    ws_bookings = wb.create_sheet("Bookings Summary")
    booking_headers = [
        'Order number', 'Internal ID', 'Buyer First Name', 'Buyer Last Name',
        'Participant Names', 'Packages', 'Participants', 'Add-ons',
        'Expected', 'Payments Received', 'Card Fee', 'Refunds', 'Net Paid', 'Balance due',
        'Status', 'Booking Date',
    ]
    _write_header(ws_bookings, booking_headers)

    money_cols = {9, 10, 11, 12, 13, 14}  # Expected … Balance due
    total_expected = 0.0
    total_received = 0.0
    total_card_fees = 0.0
    total_refunds = 0.0
    total_net_paid = 0.0
    total_balance = 0.0
    active_bookings = 0
    total_participants = 0
    cancelled_bookings = 0

    booking_row = 2
    for booking in bookings:
        package_names = []
        for bp in booking.booking_packages:
            if bp.package:
                plan = (bp.payment_plan_type or 'full').strip()
                plan_tag = 'Installment' if plan == 'deposit_installment' else 'Full'
                package_names.append(f"{bp.package.name} x{bp.quantity} ({plan_tag})")
        package_str = ', '.join(package_names) if package_names else dash

        addons_map = {}
        seen_addon_ids = set()
        for participant in booking.participants:
            for booking_addon in participant.addons:
                if booking_addon.addon and booking_addon.id not in seen_addon_ids:
                    addon_name = booking_addon.addon.name
                    addons_map[addon_name] = addons_map.get(addon_name, 0) + booking_addon.quantity
                    seen_addon_ids.add(booking_addon.id)
        for booking_addon in booking.addons:
            if booking_addon.addon and booking_addon.id not in seen_addon_ids:
                addon_name = booking_addon.addon.name
                addons_map[addon_name] = addons_map.get(addon_name, 0) + booking_addon.quantity
                seen_addon_ids.add(booking_addon.id)
        addons_str = ', '.join(
            [f"{name} x{qty}" if qty > 1 else name for name, qty in addons_map.items()]
        ) if addons_map else dash

        expected = _booking_expected(booking)
        payments_received, card_fees, refunds = _booking_payment_totals(booking)
        net_paid = round(float(booking.amount_paid) if booking.amount_paid is not None else 0.0, 2)
        is_cancelled = booking.status == 'cancelled'
        all_pax = list(booking.participants)
        active_pax = sum(
            1 for p in all_pax
            if (getattr(p, 'status', None) or 'active') != 'withdrawn'
        )
        withdrawn_pax = len(all_pax) - active_pax
        pax_count = int(booking.passenger_count or len(all_pax) or 0)

        buyer_first = booking.buyer_first_name or ''
        buyer_last = booking.buyer_last_name or ''
        if not buyer_first and not buyer_last and booking.client:
            parts = (booking.client.name or '').strip().split(None, 1)
            buyer_first = parts[0] if parts else ''
            buyer_last = parts[1] if len(parts) > 1 else ''

        if is_cancelled:
            # WeTravel 风格：取消单 Trip Price 可显示 -，仍展示已收款与退款
            expected_cell = dash
            balance_cell = dash
            participants_cell = f'0 ({pax_count} canceled)' if pax_count else '0 (canceled)'
            names_cell = _participant_names(booking, mark_cancelled=True)
            cancelled_bookings += 1
            # 退款合计仍计入汇总（便于对账）
            total_received += payments_received
            total_card_fees += card_fees
            total_refunds += refunds
        else:
            expected_cell = expected
            # 退款不计入欠款：expected − net_paid − refunds
            balance_cell = round(max(0.0, expected - net_paid - refunds), 2)
            if withdrawn_pax:
                participants_cell = f'{active_pax} (+{withdrawn_pax} withdrawn)'
            else:
                participants_cell = active_pax
            names_cell = _participant_names(booking)
            total_expected += expected
            total_received += payments_received
            total_card_fees += card_fees
            total_refunds += refunds
            total_net_paid += net_paid
            total_balance += balance_cell
            active_bookings += 1
            total_participants += active_pax

        booking_date_str = booking.created_at.strftime('%d %b %Y').upper() if booking.created_at else ''
        booking_row_data = [
            booking.order_number or f'#{booking.id}',
            booking.id,
            buyer_first or dash,
            buyer_last or dash,
            names_cell,
            package_str,
            participants_cell,
            addons_str,
            expected_cell,
            payments_received if payments_received else dash,
            card_fees if card_fees else dash,
            refunds if refunds else dash,
            net_paid,
            balance_cell,
            (booking_payment_display_status(booking) or booking.status or '').replace('_', ' ').title(),
            booking_date_str,
        ]
        row_font = cancelled_font if is_cancelled else body_font
        for col_idx, value in enumerate(booking_row_data, start=1):
            cell = ws_bookings.cell(row=booking_row, column=col_idx, value=value)
            cell.font = row_font
            cell.border = thin
            if col_idx in money_cols:
                cell.alignment = money_align
                if isinstance(value, (int, float)):
                    cell.number_format = money_format
            else:
                cell.alignment = cell_align
        booking_row += 1

    # Totals
    totals_row = booking_row + 1
    totals_label = ws_bookings.cell(
        row=totals_row, column=1,
        value=f'Totals (active {active_bookings}; cancelled {cancelled_bookings})'
    )
    totals_label.font = body_bold
    totals_label.alignment = cell_align
    cell_p = ws_bookings.cell(row=totals_row, column=7, value=total_participants)
    cell_p.font = body_bold
    cell_p.alignment = cell_align
    for col_idx, value in (
        (9, total_expected),
        (10, total_received),
        (11, total_card_fees),
        (12, total_refunds),
        (13, total_net_paid),
        (14, total_balance),
    ):
        cell = ws_bookings.cell(row=totals_row, column=col_idx, value=round(value, 2))
        cell.font = body_bold
        cell.number_format = money_format
        cell.alignment = money_align
        cell.border = thin

    # Collected Funds 摘要（参考 WeTravel 底部）
    summary_row = totals_row + 2
    ws_bookings.cell(row=summary_row, column=1, value='Collected Funds').font = body_bold
    ws_bookings.cell(row=summary_row + 1, column=1, value='Expected (active):').font = body_font
    c = ws_bookings.cell(row=summary_row + 1, column=2, value=round(total_expected, 2))
    c.number_format = money_format
    c.font = body_font
    ws_bookings.cell(row=summary_row + 2, column=1, value='Payments received:').font = body_font
    c = ws_bookings.cell(row=summary_row + 2, column=2, value=round(total_received, 2))
    c.number_format = money_format
    c.font = body_font
    ws_bookings.cell(row=summary_row + 3, column=1, value='Card fee:').font = body_font
    c = ws_bookings.cell(row=summary_row + 3, column=2, value=round(total_card_fees, 2))
    c.number_format = money_format
    c.font = body_font
    ws_bookings.cell(row=summary_row + 4, column=1, value='Refunds:').font = body_font
    c = ws_bookings.cell(row=summary_row + 4, column=2, value=round(total_refunds, 2))
    c.number_format = money_format
    c.font = body_font
    ws_bookings.cell(row=summary_row + 5, column=1, value='Net paid (active):').font = body_font
    c = ws_bookings.cell(row=summary_row + 5, column=2, value=round(total_net_paid, 2))
    c.number_format = money_format
    c.font = body_font
    ws_bookings.cell(row=summary_row + 6, column=1, value='Balance due (active):').font = body_font
    c = ws_bookings.cell(row=summary_row + 6, column=2, value=round(total_balance, 2))
    c.number_format = money_format
    c.font = body_font

    _autosize(ws_bookings)

    # ===== Sheet 4: Canceled（取消订单参与者，参考 WeTravel Canceled Information）=====
    ws_canceled = wb.create_sheet("Canceled")
    canceled_headers = [
        'Order number', 'Buyer First Name', 'Buyer Last Name',
        'Participant First Name', 'Participant Last Name', 'Email',
        'Package', 'Payments Received', 'Card Fee', 'Refunds', 'Net Paid', 'Status', 'Booking Date',
    ]
    _write_header(ws_canceled, canceled_headers)
    canceled_row = 2
    for booking in bookings:
        if booking.status != 'cancelled':
            continue
        payments_received, card_fees, refunds = _booking_payment_totals(booking)
        net_paid = round(float(booking.amount_paid) if booking.amount_paid is not None else 0.0, 2)
        buyer_first = booking.buyer_first_name or ''
        buyer_last = booking.buyer_last_name or ''
        if not buyer_first and not buyer_last and booking.client:
            parts = (booking.client.name or '').strip().split(None, 1)
            buyer_first = parts[0] if parts else ''
            buyer_last = parts[1] if len(parts) > 1 else ''
        package_names = [
            f"{bp.package.name} x{bp.quantity}" for bp in booking.booking_packages if bp.package
        ]
        package_str = ', '.join(package_names) if package_names else dash
        booking_date_str = booking.created_at.strftime('%d %b %Y').upper() if booking.created_at else ''
        order_no = booking.order_number or f'#{booking.id}'
        participants = list(booking.participants)
        if not participants:
            participants = [None]
        for participant in participants:
            if participant:
                first_name = getattr(participant, 'first_name', None) or (
                    (participant.name or '').split(None, 1)[0] if participant.name else ''
                )
                last_name = getattr(participant, 'last_name', None) or (
                    (participant.name or '').split(None, 1)[-1]
                    if participant.name and len((participant.name or '').split()) > 1 else ''
                )
                email = participant.email or booking.buyer_email or (
                    booking.client.email if booking.client else ''
                ) or dash
            else:
                first_name = last_name = dash
                email = booking.buyer_email or (booking.client.email if booking.client else '') or dash
            row_data = [
                order_no,
                buyer_first or dash,
                buyer_last or dash,
                first_name or dash,
                last_name or dash,
                email,
                package_str,
                payments_received if payments_received else dash,
                card_fees if card_fees else dash,
                refunds if refunds else dash,
                net_paid,
                'Cancelled',
                booking_date_str,
            ]
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_canceled.cell(row=canceled_row, column=col_idx, value=value)
                cell.font = cancelled_font
                cell.border = thin
                if col_idx in (8, 9, 10, 11) and isinstance(value, (int, float)):
                    cell.number_format = money_format
                    cell.alignment = money_align
                else:
                    cell.alignment = cell_align
            canceled_row += 1
    if canceled_row == 2:
        cell = ws_canceled.cell(row=2, column=1, value='No cancelled bookings')
        cell.font = body_font
    _autosize(ws_canceled)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_title = ''.join(ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in (trip.title or 'trip'))[:40]
    filename = f"Booking_Summary_{safe_title}_{id}_{datetime.now().strftime('%Y%m%d')}.xlsx"

    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"},
    )


@bp.route('/trips/<int:id>/bookings/export/data-source-url')
@login_required
def get_export_data_source_url(id):
    """获取 Excel 数据源 URL，用于调试或手动「自网站」添加连接"""
    trip = Trip.query.get_or_404(id)
    token = generate_export_token(trip.id)
    base_url = current_app.config.get('EXCEL_REFRESH_BASE_URL')
    if base_url:
        path = url_for('admin.export_bookings_csv', token=token, format='html', _external=False)
        url = base_url + path
    else:
        url = url_for('admin.export_bookings_csv', token=token, format='html', _external=True)
    return jsonify({'success': True, 'url': url})


def _get_participants_export_data(trip, include_cancelled=True):
    """提取参与者导出数据，供 CSV 和 HTML 共用。
    include_cancelled=False 时跳过已取消订单（Canceled sheet 单独列出）。
    """
    from app.payments import booking_payment_display_status

    bookings = trip.bookings.all()
    custom_questions = list(trip.questions.order_by(CustomQuestion.id).all()) if trip.questions else []
    headers = ['No', 'First Name', 'Middle Name', 'Last Name', 'Gender', 'Date of Birth', 'Registration Type',
               'Status', 'Dietary restrictions or allergies', 'Medical conditions']
    for q in custom_questions:
        headers.append(q.label)
    headers.extend([
        'Email', 'Buyer', 'Package', 'Add-ons',
        'Payment Status', 'Refunds', 'Net Paid', 'Booking Date',
    ])

    rows = []
    row_num = 1
    for booking in bookings:
        if not include_cancelled and booking.status == 'cancelled':
            continue
        package_names = [f"{bp.package.name} x{bp.quantity}" for bp in booking.booking_packages if bp.package]
        package_str = ', '.join(package_names) if package_names else '-'
        addons_map = {}
        for p in booking.participants:
            for ba in p.addons:
                if ba.addon:
                    addons_map[ba.addon.name] = addons_map.get(ba.addon.name, 0) + ba.quantity
        for ba in booking.addons:
            if ba.addon:
                addons_map[ba.addon.name] = addons_map.get(ba.addon.name, 0) + ba.quantity
        addons_str = ', '.join([f"{n} x{q}" if q > 1 else n for n, q in addons_map.items()]) if addons_map else '-'
        status_display = booking_payment_display_status(booking).replace('_', ' ').title()
        refunds_total = 0.0
        for payment in Payment.query.filter_by(booking_id=booking.id).all():
            if payment.status in ('succeeded', 'partially_refunded', 'refunded'):
                refunds_total += float(payment.refunded_amount or 0.0)
        refunds_total = round(refunds_total, 2)
        net_paid = round(float(booking.amount_paid) if booking.amount_paid is not None else 0.0, 2)
        booking_date_str = booking.created_at.strftime('%d %b %Y').upper() if booking.created_at else '-'
        buyer_name = (
            f"{booking.buyer_first_name or ''} {booking.buyer_last_name or ''}".strip()
            or (booking.client.name if booking.client else '-')
        )

        for participant in booking.participants:
            first_name = getattr(participant, 'first_name', None) or ((participant.name or '').split(None, 1)[0] if participant.name else '')
            middle_name = getattr(participant, 'middle_name', None) or ''
            last_name = getattr(participant, 'last_name', None) or ((participant.name or '').split(None, 2)[-1] if len((participant.name or '').split()) > 1 else '')
            gender = getattr(participant, 'gender', None) or ''
            dob_str = participant.dob.strftime('%Y-%m-%d') if getattr(participant, 'dob', None) else ''
            reg_type = getattr(participant, 'registration_type', None) or ''
            p_status = (getattr(participant, 'status', None) or 'active')
            if p_status == 'withdrawn':
                first_name = f'{first_name} (withdrawn)' if first_name else '(withdrawn)'
            qa = (participant.question_answers or {}) if hasattr(participant, 'question_answers') else {}
            dietary = qa.get('dietary_restrictions_or_allergies') or {}
            medical = qa.get('medical_conditions') or {}
            dietary_str = (dietary.get('details', '') or '') if dietary.get('value') == 'yes' else 'No'
            medical_str = (medical.get('details', '') or '') if medical.get('value') == 'yes' else 'No'
            row = [row_num, first_name, middle_name, last_name, gender, dob_str, reg_type, p_status, dietary_str, medical_str]
            for q in custom_questions:
                ans = qa.get(str(q.id), qa.get(q.id))
                if isinstance(ans, dict):
                    if ans.get('type') == 'file' or (
                        'original_filename' in ans and 'details' not in ans
                    ):
                        row.append(ans.get('original_filename') or ans.get('value') or '-')
                    elif 'details' in ans:
                        if ans.get('value') == 'yes' and ans.get('details'):
                            row.append(f"Yes: {ans.get('details')}")
                        else:
                            row.append(ans.get('value') or ans.get('details') or '-')
                    else:
                        row.append(ans.get('value') or '-')
                else:
                    row.append(ans or '-')
            row.extend([
                participant.email or '-',
                buyer_name,
                package_str,
                addons_str,
                status_display,
                refunds_total if refunds_total else '-',
                net_paid,
                booking_date_str,
            ])
            rows.append(row)
            row_num += 1
    return headers, rows


@bp.route('/trips/bookings/export/csv')
def export_bookings_csv():
    """导出 Participants 为 CSV 或 HTML。需通过 token 认证，无需登录。
    format=html 时返回 HTML 表格（id=participants），供 Excel Web Query 自动刷新使用。"""
    import csv
    import io
    from flask import Response
    token = request.args.get('token')
    trip_id = verify_export_token(token)
    if not trip_id:
        return jsonify({'error': 'Invalid or expired token'}), 403

    trip = Trip.query.get(trip_id)
    if not trip:
        return jsonify({'error': 'Trip not found'}), 404

    headers, rows = _get_participants_export_data(trip)

    if request.args.get('format') == 'html':
        # HTML 表格供 Excel Web Query 抓取
        from html import escape as html_escape
        html_parts = ['<!DOCTYPE html><html><head><meta charset="utf-8"></head><body>']
        html_parts.append('<table id="participants"><thead><tr>')
        for h in headers:
            html_parts.append(f'<th>{html_escape(str(h))}</th>')
        html_parts.append('</tr></thead><tbody>')
        for row in rows:
            html_parts.append('<tr>')
            for cell in row:
                html_parts.append(f'<td>{html_escape(str(cell))}</td>')
            html_parts.append('</tr>')
        html_parts.append('</tbody></table></body></html>')
        html_content = ''.join(html_parts)
        return Response(
            html_content,
            mimetype='text/html; charset=utf-8',
            headers={
                'Cache-Control': 'no-cache, no-store, must-revalidate',
                'Pragma': 'no-cache'
            }
        )

    # 默认返回 CSV
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    output.seek(0)
    csv_content = '\ufeff' + output.getvalue()  # BOM for Excel UTF-8
    return Response(
        csv_content,
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': 'inline; filename="participants.csv"',
            'Cache-Control': 'no-cache, no-store, must-revalidate',
            'Pragma': 'no-cache'
        }
    )


@bp.route('/trips/<int:id>/bookings/add', methods=['POST'])
@login_required
def add_participant(id):
    trip = Trip.query.get_or_404(id)
    
    # Check for JSON payload (from Multi-step Modal)
    if request.is_json:
        data = request.get_json()
        try:
            # 1. Extract Buyer Info
            buyer_info = data.get('buyer', {})
            email = buyer_info.get('email')
            if not email:
                return jsonify({'success': False, 'message': 'Buyer email is required'}), 400
            
            # Extract buyer name (support both full name and first_name/last_name)
            buyer_first_name = buyer_info.get('first_name', '')
            buyer_last_name = buyer_info.get('last_name', '')
            buyer_name = buyer_info.get('name', '')
            if not buyer_name and (buyer_first_name or buyer_last_name):
                buyer_name = f"{buyer_first_name} {buyer_last_name}".strip()
            if not buyer_name:
                buyer_name = email.split('@')[0]  # Fallback to email username
                
            client = Client.query.filter_by(email=email).first()
            if not client:
                client = Client(
                    name=buyer_name,
                    first_name=buyer_first_name,
                    last_name=buyer_last_name,
                    email=email,
                    phone=buyer_info.get('phone'),
                    address=buyer_info.get('address'),
                    city=buyer_info.get('city'),
                    state=buyer_info.get('state'),
                    zip_code=buyer_info.get('zip_code'),
                    country=buyer_info.get('country')
                )
                db.session.add(client)
                db.session.flush()
            else:
                # Update existing client with new buyer info if provided
                if buyer_first_name:
                    client.first_name = buyer_first_name
                if buyer_last_name:
                    client.last_name = buyer_last_name
                if buyer_name:
                    client.name = buyer_name
                if buyer_info.get('phone'):
                    client.phone = buyer_info.get('phone')
                if buyer_info.get('address'):
                    client.address = buyer_info.get('address')
                if buyer_info.get('city'):
                    client.city = buyer_info.get('city')
                if buyer_info.get('state'):
                    client.state = buyer_info.get('state')
                if buyer_info.get('zip_code'):
                    client.zip_code = buyer_info.get('zip_code')
                if buyer_info.get('country'):
                    client.country = buyer_info.get('country')
            
            # 2. Create Booking
            packages_data = data.get('packages', [])  # New format: list of {package_id, quantity, payment_plan_type}
            participants_data = data.get('participants', [])
            payment_info = data.get('payment', {})
            
            # Calculate total amount paid
            try:
                amount_paid = float(payment_info.get('amount_paid', 0) or 0)
            except (ValueError, TypeError):
                amount_paid = 0.0
            
            # Validate amount_paid is not negative
            if amount_paid < 0:
                return jsonify({'success': False, 'message': 'Amount paid cannot be negative'}), 400
            
            # Calculate total passenger count from packages
            total_passengers = sum(pkg.get('quantity', 1) for pkg in packages_data) if packages_data else len(participants_data)
            
            # Validate packages_data is not empty
            if not packages_data or len(packages_data) == 0:
                return jsonify({'success': False, 'message': 'At least one package must be selected'}), 400
            
            # Validate participants_data matches total_passengers
            if len(participants_data) != total_passengers:
                return jsonify({'success': False, 'message': f'Number of participants ({len(participants_data)}) does not match package quantities ({total_passengers})'}), 400
            
            # Extract buyer custom info if provided
            buyer_custom_info = buyer_info.get('custom_info')
            if isinstance(buyer_custom_info, str):
                try:
                    buyer_custom_info = json.loads(buyer_custom_info)
                except:
                    buyer_custom_info = None
            
            booking = Booking(
                trip_id=trip.id,
                client=client,
                passenger_count=total_passengers,
                amount_paid=amount_paid,
                status=payment_info.get('status', 'deposit_paid'),
                # Buyer Info fields
                buyer_first_name=buyer_first_name or client.first_name,
                buyer_last_name=buyer_last_name or client.last_name,
                buyer_email=email,
                buyer_phone=buyer_info.get('phone') or client.phone,
                buyer_address=buyer_info.get('address') or client.address,
                buyer_city=buyer_info.get('city') or client.city,
                buyer_state=buyer_info.get('state') or client.state,
                buyer_zip_code=buyer_info.get('zip_code') or client.zip_code,
                buyer_country=buyer_info.get('country') or client.country,
                buyer_emergency_contact_name=buyer_info.get('emergency_contact_name'),
                buyer_emergency_contact_phone=buyer_info.get('emergency_contact_phone'),
                buyer_emergency_contact_email=buyer_info.get('emergency_contact_email'),
                buyer_emergency_contact_relationship=buyer_info.get('emergency_contact_relationship'),
                buyer_home_phone=buyer_info.get('home_phone'),
                buyer_work_phone=buyer_info.get('work_phone'),
                buyer_custom_info=buyer_custom_info
            )
            db.session.add(booking)
            db.session.flush()

            from app.order_numbers import assign_order_number
            assign_order_number(booking, trip=trip)
            
            # 2a. Create BookingPackages
            booking_packages_created = []
            for pkg_data in packages_data:
                package_id = pkg_data.get('package_id')
                quantity = int(pkg_data.get('quantity', 1))
                payment_plan_type = pkg_data.get('payment_plan_type', 'full')
                
                if package_id and quantity > 0:
                    package = TripPackage.query.get(package_id)
                    booking_package = BookingPackage(
                        booking_id=booking.id,
                        package_id=package_id,
                        quantity=quantity,
                        payment_plan_type=payment_plan_type,
                        amount_paid=0.0,  # Will be updated based on payment
                        status='pending',
                        unit_price=(
                            float(package.price)
                            if package and package.price is not None
                            else 0.0
                        ),
                    )
                    db.session.add(booking_package)
                    booking_packages_created.append((booking_package, package))
            db.session.flush()

            # Create installment rows for deposit_installment packages (same as customer path)
            from app.routes import create_installment_payments
            for booking_package, package in booking_packages_created:
                if (
                    booking_package.payment_plan_type == 'deposit_installment'
                    and package
                    and package.payment_plan_config
                    and package.payment_plan_config.get('enabled')
                ):
                    create_installment_payments(booking, booking_package, package.payment_plan_config)
            
            # 3. Create Participants & Add-ons
            for p_data in participants_data:
                # Use first_name and last_name if available, otherwise fallback to name
                participant_name = p_data.get('name')
                if p_data.get('first_name') or p_data.get('last_name'):
                    participant_name = f"{p_data.get('first_name', '')} {p_data.get('last_name', '')}".strip() or participant_name
                
                question_answers = p_data.get('question_answers', {}) or {}
                dob_val = None
                if p_data.get('dob'):
                    try:
                        dob_val = datetime.strptime(str(p_data['dob']), '%Y-%m-%d').date()
                    except (ValueError, TypeError):
                        pass
                participant = BookingParticipant(
                    booking_id=booking.id,
                    name=participant_name,
                    email=p_data.get('email'),
                    first_name=p_data.get('first_name'),
                    middle_name=p_data.get('middle_name'),
                    last_name=p_data.get('last_name'),
                    gender=p_data.get('gender'),
                    dob=dob_val,
                    registration_type=p_data.get('registration_type'),
                    question_answers=question_answers if question_answers else None,
                )
                db.session.add(participant)
                db.session.flush()
                
                # Handle Add-ons for this participant
                # payload format: "addons": {"1": 2, "3": 1} (addon_id: quantity)
                addons_selection = p_data.get('addons', {}) 
                for addon_id, qty in addons_selection.items():
                    if int(qty) > 0:
                        ba = BookingAddOn(
                            booking_id=booking.id,
                            participant_id=participant.id,
                            addon_id=int(addon_id),
                            quantity=int(qty)
                        )
                        db.session.add(ba)
            
            db.session.commit()
            return jsonify({'success': True, 'message': 'Participants added successfully'})
            
        except Exception as e:
            db.session.rollback()
            # print(e) # Debug
            return jsonify({'success': False, 'message': str(e)}), 500

    return jsonify({'success': False, 'message': 'Invalid request format'}), 400


@bp.route('/trips/<int:trip_id>/bookings/<int:booking_id>', methods=['GET', 'POST'])
@login_required
def manage_booking(trip_id, booking_id):
    """查看和编辑单个预订"""
    trip = Trip.query.get_or_404(trip_id)
    booking = Booking.query.get_or_404(booking_id)
    
    # 确保 booking 属于这个 trip
    if booking.trip_id != trip.id:
        if request.args.get('format') == 'json' or request.is_json or request.method == 'POST':
            return jsonify({'success': False, 'message': 'Booking does not belong to this trip'}), 400
        flash('Booking does not belong to this trip', 'error')
        return redirect(url_for('admin.manage_trip', id=trip_id))
    
    form = EditBookingForm()
    
    if request.method == 'GET':
        # 返回 JSON 数据用于模态框显示
        if request.args.get('format') == 'json':
            # 获取参与者信息
            participants = []
            for p in booking.participants:
                participant_addons = []
                for ba in p.addons:
                    if ba.addon:
                        participant_addons.append({
                            'id': ba.addon.id,
                            'name': ba.addon.name,
                            'price': booking_addon_unit_price(ba),
                            'quantity': ba.quantity
                        })
                # Get question_answers if stored (currently not in model, but prepare for future)
                question_answers = {}
                if hasattr(p, 'question_answers') and p.question_answers:
                    import json
                    if isinstance(p.question_answers, str):
                        try:
                            question_answers = json.loads(p.question_answers)
                        except:
                            question_answers = {}
                    elif isinstance(p.question_answers, dict):
                        question_answers = p.question_answers
                
                name_parts = (p.name or '').strip().split()
                first_name = getattr(p, 'first_name', None) or (name_parts[0] if name_parts else '')
                last_name = getattr(p, 'last_name', None) or (name_parts[-1] if len(name_parts) > 1 else '')
                middle_name = getattr(p, 'middle_name', None) or (name_parts[1] if len(name_parts) > 2 else '')
                participants.append({
                    'id': p.id,
                    'name': p.name,
                    'email': p.email,
                    'phone': p.phone,
                    'first_name': first_name,
                    'middle_name': middle_name,
                    'last_name': last_name,
                    'gender': getattr(p, 'gender', None) or '',
                    'dob': (lambda d: d.strftime('%Y-%m-%d') if d else '')(getattr(p, 'dob', None)),
                    'registration_type': getattr(p, 'registration_type', None) or '',
                    'question_answers': question_answers,
                    'addons': participant_addons,
                    'status': getattr(p, 'status', None) or 'active',
                    'is_withdrawn': (getattr(p, 'status', None) or 'active') == 'withdrawn',
                })
            
            # 获取所有附加项（包括直接关联到 booking 的，不通过 participant）
            from app.addon_admin import serialize_booking_addon
            all_addons = []
            addons_total = 0.0
            seen_addon_ids = set()
            
            # 方法1：通过 participant.addons 获取
            for p in booking.participants:
                for ba in p.addons:
                    if ba.addon and ba.id not in seen_addon_ids:
                        row = serialize_booking_addon(ba)
                        addons_total += float(row['subtotal'])
                        all_addons.append(row)
                        seen_addon_ids.add(ba.id)
            
            # 方法2：通过 booking.addons 获取（直接关联的）
            for ba in booking.addons:
                if ba.addon and ba.id not in seen_addon_ids:
                    row = serialize_booking_addon(ba)
                    addons_total += float(row['subtotal'])
                    all_addons.append(row)
                    seen_addon_ids.add(ba.id)
            
            # 计算套餐金额
            packages_total = 0.0
            packages_data = []
            has_packages = False
            
            for bp in booking.booking_packages:
                if bp.package:
                    package_price = booking_package_unit_price(bp)
                    quantity = int(bp.quantity) if bp.quantity is not None else 1
                    subtotal = package_price * quantity
                    packages_total += subtotal
                    has_packages = True
                    
                    # 获取分期付款配置
                    payment_plan_config = None
                    if bp.payment_plan_type == 'deposit_installment' and bp.package.payment_plan_config:
                        payment_plan_config = bp.package.payment_plan_config
                    
                    packages_data.append({
                        'id': bp.package.id,
                        'name': bp.package.name,
                        'price': package_price,
                        'quantity': quantity,
                        'subtotal': subtotal,
                        'payment_plan_type': bp.payment_plan_type,
                        'payment_plan_config': payment_plan_config
                    })
            
            # 计算应付金额（扣除折扣）
            discount_amount = float(booking.discount_amount) if booking.discount_amount else 0.0
            expected_amount = max(0.0, packages_total + addons_total - discount_amount)
            
            # Fallback for legacy bookings
            if not has_packages:
                expected_amount = float(booking.amount_paid) if booking.amount_paid is not None else 0.0
            
            # -------------------------------------------------------------------------
            # 支付数据流（与 routes 中 Stripe 回调一致）：
            # - Payment.amount = 客户该笔实扣（美元，含卡手续费）
            # - Payment.fee_cents = 该笔卡手续费（分）；若无则用 base_amount_cents 反推或 0
            # - Booking.amount_paid = 已付基础金额合计（不含手续费），支付成功时在 routes 中累加
            # - expected_amount = 订单应付（套餐+附加-折扣），与 Payment 无关
            # - payment_summary = 从 Payment 表汇总；amount_paid_display = 有支付时用汇总实收，否则用 booking.amount_paid
            # -------------------------------------------------------------------------
            payments = []
            total_charged = 0.0
            total_fee_dollars = 0.0
            net_received = 0.0
            total_refunded = 0.0
            from app.payments import (
                payment_charged_amount,
                payment_base_amount,
                payment_fee_amount,
                payment_refundable_remaining,
                payment_max_refund,
                payment_refunded_clamped,
                booking_deposit_reserved,
                booking_payment_display_status,
                booking_payment_type_display,
                booking_balance_due,
                payment_step_label,
                payment_is_payoff,
                booking_has_payoff,
            )
            settled_via_payoff = False
            for payment in (
                Payment.query.options(joinedload(Payment.installment_payment))
                .filter_by(booking_id=booking.id)
                .order_by(Payment.paid_at.asc(), Payment.created_at.asc(), Payment.id.asc())
                .all()
            ):
                amt = payment_charged_amount(payment)
                fee_dollars = payment_fee_amount(payment)
                base_dollars = payment_base_amount(payment)
                refunded = payment_refunded_clamped(payment)
                remaining = payment_refundable_remaining(payment)
                # 实收：基础金额 − 已退（已为基础口径）
                net_from_payment = round(max(0.0, base_dollars - refunded), 2)
                is_payoff = payment_is_payoff(payment)
                if is_payoff and payment.status in ('succeeded', 'partially_refunded', 'refunded'):
                    settled_via_payoff = True

                payments.append({
                    'id': payment.id,
                    'amount': amt,
                    'base_amount': base_dollars,
                    'fee_amount': fee_dollars,
                    'status': payment.status,
                    'type_label': payment_step_label(payment),
                    'is_payoff': is_payoff,
                    'paid_at': payment.paid_at.strftime('%Y-%m-%d %H:%M') if payment.paid_at else None,
                    'created_at': payment.created_at.strftime('%Y-%m-%d %H:%M') if payment.created_at else None,
                    'stripe_payment_intent_id': payment.stripe_payment_intent_id,
                    'fee_cents': payment.fee_cents,
                    'funding': payment.funding,
                    'brand': payment.brand,
                    'payment_method_type': payment.payment_method_type,
                    'refunded_amount': refunded,
                    'refundable_remaining': remaining,
                    # 默认不含定金的可退上限；勾选 include_deposit 后用 refundable_remaining
                    'max_refund_without_deposit': payment_max_refund(
                        payment, booking, include_deposit=False, deposit_hint=None
                    ),
                    'refund_reason': payment.refund_reason or '',
                    'can_refund': payment.status in ('succeeded', 'partially_refunded') and remaining > 0.001,
                })
                if payment.status in ('succeeded', 'partially_refunded', 'refunded'):
                    total_charged += amt
                    total_fee_dollars += fee_dollars
                    net_received += net_from_payment
                    total_refunded += refunded
            net_received = round(net_received, 2)
            total_fee_dollars = round(total_fee_dollars, 2)
            total_charged = round(total_charged, 2)
            total_refunded = round(total_refunded, 2)
            payment_summary = {
                'total_charged': total_charged,
                'card_fee': total_fee_dollars,
                'refunded': total_refunded,
                'net_received': net_received,
            }
            # 有支付记录时用汇总实收作为“已付金额”显示（与 Payment 表一致，避免 DB 存错）
            amount_paid_display = round(net_received if total_charged > 0 else (float(booking.amount_paid) if booking.amount_paid else 0.0), 2)

            # 定金参考：仅统计分期套餐的 deposit，多包求和（混单勿用全款包的 config）
            deposit_hint = 0.0
            has_inst_deposit = False
            for bp in booking.booking_packages:
                if (bp.payment_plan_type or 'full') != 'deposit_installment':
                    continue
                cfg = (bp.package.payment_plan_config if bp.package else None) or {}
                dep = cfg.get('deposit_amount') or cfg.get('deposit')
                if dep is None:
                    continue
                try:
                    deposit_hint += float(dep) * (int(bp.quantity) if bp.quantity else 1)
                    has_inst_deposit = True
                except (TypeError, ValueError):
                    pass
            deposit_hint = deposit_hint if has_inst_deposit else None
            deposit_reserved = booking_deposit_reserved(booking, deposit_hint=deposit_hint)

            # 获取分期付款记录
            from app.payments import installment_display_label
            inst_rows = (
                InstallmentPayment.query.filter_by(booking_id=booking.id)
                .order_by(InstallmentPayment.installment_number)
                .all()
            )
            post_deposit_count = sum(1 for i in inst_rows if (i.installment_number or 0) > 0)
            installments = []
            for inst in inst_rows:
                installments.append({
                    'id': inst.id,
                    'installment_number': inst.installment_number,
                    'label': installment_display_label(
                        inst, post_deposit_count=post_deposit_count
                    ),
                    'amount': float(inst.amount) if inst.amount else 0.0,
                    'due_date': inst.due_date.strftime('%Y-%m-%d') if inst.due_date else None,
                    'status': inst.status,
                    'paid_at': inst.paid_at.strftime('%Y-%m-%d %H:%M') if inst.paid_at else None
                })
            
            # 获取折扣码信息（如果有）
            discount_info = None
            if booking.discount_code_id and booking.discount_code:
                discount_info = {
                    'code': booking.discount_code.code,
                    'type': booking.discount_code.type,
                    'value': booking.discount_code.amount,
                    'discount_amount': float(booking.discount_amount) if booking.discount_amount else 0.0
                }

            pending_amount = (
                booking_balance_due(booking, expected=expected_amount)
                if booking.status != 'cancelled'
                else 0.0
            )
            
            return jsonify({
                'success': True,
                'booking': {
                    'id': booking.id,
                    'order_number': booking.order_number,
                    'order_seq': booking.order_seq,
                    # Buyer 详细信息
                    'buyer': {
                        'first_name': booking.buyer_first_name or '',
                        'last_name': booking.buyer_last_name or '',
                        'name': f"{booking.buyer_first_name or ''} {booking.buyer_last_name or ''}".strip() or (booking.client.name if booking.client else ''),
                        'email': booking.buyer_email or (booking.client.email if booking.client else ''),
                        'phone': booking.buyer_phone or (booking.client.phone if booking.client else ''),
                        'address': booking.buyer_address or '',
                        'city': booking.buyer_city or '',
                        'state': booking.buyer_state or '',
                        'zip_code': booking.buyer_zip_code or '',
                        'country': booking.buyer_country or '',
                        'emergency_contact_name': booking.buyer_emergency_contact_name or '',
                        'emergency_contact_phone': booking.buyer_emergency_contact_phone or '',
                        'emergency_contact_email': booking.buyer_emergency_contact_email or '',
                        'emergency_contact_relationship': booking.buyer_emergency_contact_relationship or ''
                    },
                    # 兼容旧的 client 字段
                    'client': {
                        'name': f"{booking.buyer_first_name or ''} {booking.buyer_last_name or ''}".strip() or (booking.client.name if booking.client else ''),
                        'email': booking.buyer_email or (booking.client.email if booking.client else ''),
                        'phone': booking.buyer_phone or (booking.client.phone if booking.client else '')
                    },
                    # 套餐信息
                    'packages': packages_data,
                    'packages_total': packages_total,
                    # 附加项信息
                    'addons': all_addons,
                    'addons_total': addons_total,
                    # 折扣信息
                    'discount': discount_info,
                    # 金额信息（amount_paid 与 display 同源，避免 Save 冲掉账本）
                    'status': booking.status,
                    'payment_status': booking_payment_display_status(booking),
                    'payment_type': booking_payment_type_display(booking),
                    'amount_paid': amount_paid_display,
                    'amount_paid_display': amount_paid_display,
                    'expected_amount': expected_amount,
                    'pending_amount': pending_amount,
                    'total_refunded': total_refunded,
                    'payment_summary': payment_summary,
                    'deposit_hint': deposit_hint,
                    'deposit_reserved': deposit_reserved,
                    'fees_non_refundable': True,
                    # 其他信息
                    'passenger_count': booking.passenger_count,
                    'special_requests': booking.special_requests or '',
                    'created_at': booking.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                    # 参与者信息
                    'participants': participants,
                    # 支付历史
                    'payments': payments,
                    # 分期付款记录
                    'installments': installments,
                    'settled_via_payoff': settled_via_payoff or booking_has_payoff(booking),
                    'auto_pay_enabled': bool(getattr(booking, 'auto_pay_enabled', False)),
                    'auto_pay_opt_in': bool(getattr(booking, 'auto_pay_opt_in', False)),
                    'auto_pay_payment_method_id': getattr(booking, 'auto_pay_payment_method_id', None),
                    'auto_pay_has_method': bool(getattr(booking, 'auto_pay_payment_method_id', None)),
                }
            })
    
    # POST 请求：更新 booking
    if request.method == 'POST':
        # Handle form data (from FormData)
        if request.form:
            prev_status = booking.status
            new_status = request.form.get('status', booking.status)
            # 取消订单请走 /cancel；Save Changes 不允许把状态设为 cancelled
            if new_status == 'cancelled' and prev_status != 'cancelled':
                return jsonify({
                    'success': False,
                    'message': 'Use Cancel order to cancel this booking (not Payment Status).',
                }), 400
            if prev_status == 'cancelled':
                # 已取消：保持 cancelled，仍可改备注/参与者等；忽略付款状态下拉
                new_status = 'cancelled'
            elif new_status not in ('pending', 'processing', 'deposit_paid', 'fully_paid'):
                return jsonify({
                    'success': False,
                    'message': 'Invalid payment status.',
                }), 400
            booking.status = new_status
            _ap = request.form.get('amount_paid')
            if _ap is not None and _ap != '':
                try:
                    booking.amount_paid = float(_ap)
                except (TypeError, ValueError):
                    pass
            booking.special_requests = request.form.get('special_requests', booking.special_requests)
            
            # Update participants (matching builder structure: first_name, last_name, question_answers)
            participants_data = request.form.get('participants')
            if participants_data:
                try:
                    import json
                    participants_list = json.loads(participants_data)
                    for p_data in participants_list:
                        participant = BookingParticipant.query.get(p_data.get('id'))
                        if participant and participant.booking_id == booking.id:
                            first_name = p_data.get('first_name', '')
                            middle_name = p_data.get('middle_name', '')
                            last_name = p_data.get('last_name', '')
                            participant.first_name = first_name or None
                            participant.middle_name = middle_name or None
                            participant.last_name = last_name or None
                            participant.name = ' '.join(filter(None, [first_name, middle_name, last_name])).strip() or participant.name
                            participant.email = p_data.get('email', participant.email)
                            participant.gender = p_data.get('gender') or None
                            participant.registration_type = p_data.get('registration_type') or None
                            if p_data.get('dob'):
                                try:
                                    participant.dob = datetime.strptime(str(p_data['dob']), '%Y-%m-%d').date()
                                except (ValueError, TypeError):
                                    pass
                            question_answers = p_data.get('question_answers', {})
                            if question_answers is not None:
                                participant.question_answers = question_answers
                except (json.JSONDecodeError, KeyError) as e:
                    # Log error but don't fail the whole request
                    print(f"Error updating participants: {e}")

            db.session.commit()
            return jsonify({'success': True, 'message': 'Booking updated successfully'})
        
        # Handle form validation (if using WTForms)
        if form.validate_on_submit():
            prev_status = booking.status
            new_status = form.status.data
            if new_status == 'cancelled' and prev_status != 'cancelled':
                if request.is_json or request.headers.get('Content-Type') == 'application/json':
                    return jsonify({
                        'success': False,
                        'message': 'Use Cancel order to cancel this booking (not Payment Status).',
                    }), 400
                flash('Use Cancel order to cancel this booking.', 'error')
                return redirect(url_for('admin.manage_trip', id=trip_id))
            if prev_status == 'cancelled':
                new_status = 'cancelled'
            booking.status = new_status
            booking.amount_paid = form.amount_paid.data
            booking.special_requests = form.special_requests.data
            db.session.commit()
            
            if request.is_json or request.headers.get('Content-Type') == 'application/json':
                return jsonify({'success': True, 'message': 'Booking updated successfully'})
            else:
                flash('Booking updated successfully', 'success')
                return redirect(url_for('admin.manage_trip', id=trip_id))
    
    # 如果不是 JSON 请求，返回表单页面（备用）
    form.status.data = booking.status
    form.amount_paid.data = booking.amount_paid
    form.special_requests.data = booking.special_requests
    
    return render_template('admin/trips/manage_booking.html', 
                         trip=trip, booking=booking, form=form)


@bp.route('/trips/<int:trip_id>/bookings/<int:booking_id>/receipt')
@login_required
def generate_receipt(trip_id, booking_id):
    """生成预订收据 PDF（默认下载）；?payment_id= 指定当笔；?format=html 查看网页版。"""
    from flask import make_response
    from app.receipt_pdf import build_booking_receipt_pdf
    from app.routes import _booking_receipt_context

    trip = Trip.query.get_or_404(trip_id)
    booking = Booking.query.get_or_404(booking_id)

    if booking.trip_id != trip.id:
        flash('Booking does not belong to this trip', 'error')
        return redirect(url_for('admin.manage_trip', id=trip_id))

    payment_id = request.args.get('payment_id', type=int)
    ctx = _booking_receipt_context(booking, payment_id=payment_id)
    if not ctx:
        # Manage 用 fetch 下载；勿 flash+redirect（否则会把 HTML 当下 PDF）
        msg = (
            'Invalid payment for this booking receipt'
            if payment_id is not None
            else 'Unable to build receipt for this booking'
        )
        return jsonify({'success': False, 'message': msg}), 400

    if request.args.get('format') == 'html':
        return render_template('booking/receipt.html', **ctx)

    pdf_bytes = build_booking_receipt_pdf(ctx)
    order_label = getattr(booking, 'order_number', None) or booking.id
    focus_pid = ctx.get('receipt_payment_id') or payment_id
    if focus_pid:
        filename = f'NHTours-Order-{order_label}-Pay-{focus_pid}.pdf'
    else:
        filename = f'NHTours-Order-{order_label}.pdf'
    response = make_response(pdf_bytes)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@bp.route('/trips/<int:trip_id>/bookings/<int:booking_id>/reconcile-ledger')
@login_required
def reconcile_booking_ledger_route(trip_id, booking_id):
    """只读核对：Booking.amount_paid vs Σ(payment base − refunded)；?stripe=1 时对比 Stripe。"""
    from app.payments import reconcile_booking_ledger

    trip = Trip.query.get_or_404(trip_id)
    booking = Booking.query.get_or_404(booking_id)
    if booking.trip_id != trip.id:
        return jsonify({'success': False, 'message': 'Booking does not belong to this trip'}), 400
    check_stripe = (request.args.get('stripe') or '').strip().lower() in ('1', 'true', 'yes')
    result = reconcile_booking_ledger(booking, check_stripe=check_stripe)
    if not result.get('ok'):
        try:
            from app.ledger_alerts import notify_ledger_mismatch
            notify_ledger_mismatch(booking, result, source='reconcile_api')
        except Exception as e:
            current_app.logger.warning('ledger alert after reconcile failed: %s', e)
    return jsonify({'success': True, 'reconcile': result})


@bp.route('/trips/<int:trip_id>/bookings/<int:booking_id>/cancel', methods=['POST'])
@login_required
def cancel_booking_order(trip_id, booking_id):
    """取消订单（不退款）。未付分期一并取消。"""
    from app.payments import cancel_unpaid_installments

    trip = Trip.query.get_or_404(trip_id)
    booking = Booking.query.get_or_404(booking_id)
    if booking.trip_id != trip.id:
        return jsonify({'success': False, 'message': 'Booking does not belong to this trip'}), 400

    if booking.status == 'cancelled':
        return jsonify({'success': True, 'message': 'Order is already cancelled', 'status': 'cancelled'})

    try:
        booking.status = 'cancelled'
        cancel_unpaid_installments(booking)
        try:
            from app.payments import void_stale_pending_payments
            void_stale_pending_payments(booking)
        except Exception:
            pass
        for bp in booking.booking_packages.all():
            if (bp.status or '') not in ('cancelled',):
                bp.status = 'cancelled'
        try:
            from app.auto_pay import disable_auto_pay
            disable_auto_pay(booking, source='admin_cancel')
        except Exception:
            pass
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error cancelling booking {booking_id}: {e}')
        return jsonify({'success': False, 'message': 'Failed to cancel order'}), 500

    return jsonify({
        'success': True,
        'message': 'Order cancelled',
        'status': 'cancelled',
        'order_number': booking.order_number,
    })


@bp.route('/trips/<int:id>/financials')
@login_required
def get_trip_financials(id):
    """获取行程的财务数据（用于 AJAX 更新）"""
    from app.payments import booking_balance_due, booking_refunded_total

    trip = Trip.query.get_or_404(id)
    bookings = trip.bookings.all() if trip.bookings else []
    
    # Financial Calculations (same logic as manage_trip)
    # amount_paid 是客户实际支付的基础金额（不含 Stripe 手续费）
    total_paid = sum(b.amount_paid or 0.0 for b in bookings)
    total_gross = 0.0
    total_discount = 0.0
    total_expected = 0.0
    total_pending = 0.0
    total_refunded = 0.0
    
    for b in bookings:
        booking_gross = 0.0
        has_packages = False
        seen_addon_ids = set()
        
        # Calculate expected amount from BookingPackages
        for bp in b.booking_packages:
            if bp.package:  # Check if package exists
                package_price = booking_package_unit_price(bp)
                quantity = int(bp.quantity) if bp.quantity is not None else 1
                booking_gross += package_price * quantity
                has_packages = True
        
        # Add add-ons prices (通过 participant.addons)
        for participant in b.participants:
            for booking_addon in participant.addons:
                if booking_addon.addon and booking_addon.id not in seen_addon_ids:
                    addon_price = booking_addon_unit_price(booking_addon)
                    quantity = int(booking_addon.quantity) if booking_addon.quantity is not None else 0
                    booking_gross += addon_price * quantity
                    seen_addon_ids.add(booking_addon.id)
        
        # Add add-ons prices (直接通过 booking.addons)
        for booking_addon in b.addons:
            if booking_addon.addon and booking_addon.id not in seen_addon_ids:
                addon_price = booking_addon_unit_price(booking_addon)
                quantity = int(booking_addon.quantity) if booking_addon.quantity is not None else 0
                booking_gross += addon_price * quantity
                seen_addon_ids.add(booking_addon.id)
        
        # 折扣金额
        discount = float(b.discount_amount) if b.discount_amount else 0.0
        booking_expected = max(0.0, booking_gross - discount)
        
        # Fallback for legacy bookings without BookingPackages
        if not has_packages:
            booking_gross = float(b.amount_paid) if b.amount_paid is not None else 0.0
            booking_expected = booking_gross
            discount = 0.0
        
        total_gross += booking_gross
        total_discount += discount
        total_expected += booking_expected
        total_refunded += booking_refunded_total(b)
        due = booking_balance_due(b, expected=booking_expected)
        if due is not None:
            total_pending += due
    
    return jsonify({
        'success': True,
        'financials': {
            'gross_amount': round(total_gross, 2),
            'total_discount': round(total_discount, 2),
            'expected_amount': round(total_expected, 2),
            'amount_paid': round(total_paid, 2),
            'total_refunded': round(total_refunded, 2),
            'amount_pending': round(total_pending, 2)
        }
    })


@bp.route('/trips/<int:trip_id>/bookings/<int:booking_id>/refund', methods=['POST'])
@admin_required
def refund_booking(trip_id, booking_id):
    """
    处理预订退款：
    - 部分退款：传 payment_id + 手填金额（只退该笔）
    - 全额退款：full_refund=true，对该订单每笔可退 Payment 退光剩余额（各原路返回）
    - $0 时允许仅 cancel_booking
    """
    trip = Trip.query.get_or_404(trip_id)
    booking = Booking.query.get_or_404(booking_id)

    if booking.trip_id != trip.id:
        return jsonify({'success': False, 'message': 'Booking does not belong to this trip'}), 400

    data = request.get_json() or {}
    try:
        refund_amount = float(data.get('amount', 0) or 0)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'message': 'Invalid refund amount'}), 400

    reason = (data.get('reason') or '').strip()
    manual_only = bool(data.get('manual_only'))
    cancel_booking = bool(data.get('cancel_booking'))
    full_refund = bool(data.get('full_refund'))
    payment_id_raw = data.get('payment_id')

    if not reason:
        return jsonify({'success': False, 'message': 'Refund reason is required'}), 400

    # --- $0 / 仅取消：无退款金额时只改订单状态 ---
    if refund_amount <= 0.001 and not full_refund:
        if not cancel_booking:
            return jsonify({
                'success': False,
                'message': 'Invalid refund amount (or check “Cancel booking” for $0 / no-refund cancel)'
            }), 400
        try:
            booking.status = 'cancelled'
            if float(booking.amount_paid or 0) <= 0.001:
                booking.amount_paid = 0.0
            from app.payments import cancel_unpaid_installments, void_stale_pending_payments
            cancel_unpaid_installments(booking)
            try:
                void_stale_pending_payments(booking)
            except Exception:
                pass
            for bp in booking.booking_packages.all():
                if (bp.status or '') not in ('cancelled',):
                    bp.status = 'cancelled'
            try:
                from app.auto_pay import disable_auto_pay
                disable_auto_pay(booking, source='admin_cancel')
            except Exception:
                pass
            db.session.commit()
            return jsonify({
                'success': True,
                'message': 'Booking cancelled (no refund amount)',
                'stripe_refund_id': None,
                'manual_only': True,
                'new_amount_paid': float(booking.amount_paid or 0),
                'booking_status': booking.status,
                'payment_refunded_amount': None,
            })
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error cancelling $0 booking: {str(e)}")
            return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500

    from app.payments import (
        process_refund,
        apply_refund_to_ledger,
        payment_refundable_remaining,
        extract_stripe_charge_id,
        retrieve_payment_intent,
    )

    def _refundable_payments():
        rows = (
            Payment.query.filter(
                Payment.booking_id == booking.id,
                Payment.status.in_(('succeeded', 'partially_refunded')),
            )
            .order_by(Payment.created_at.asc(), Payment.id.asc())
            .all()
        )
        out = []
        for p in rows:
            avail = payment_refundable_remaining(p)
            if avail > 0.001:
                out.append((p, avail))
        return out

    def _stripe_refund_payment(payment, slice_amount):
        pi = payment.stripe_payment_intent_id
        if not pi or not str(pi).startswith('pi_'):
            return None, (
                f'Payment #{payment.id} has no Stripe PaymentIntent. '
                'Use “Manual only” to record a local refund, or refund in Stripe Dashboard.'
            )
        if not payment.stripe_charge_id:
            try:
                intent = retrieve_payment_intent(pi)
                charge_id = extract_stripe_charge_id(intent)
                if charge_id:
                    payment.stripe_charge_id = charge_id
            except Exception as e:
                current_app.logger.warning(f'Could not fetch charge for {pi}: {e}')
        stripe_refund, err = process_refund(pi, slice_amount, reason)
        if err or not stripe_refund:
            return None, f'Stripe refund failed on payment #{payment.id}: {err or "unknown error"}'
        return getattr(stripe_refund, 'id', None), None

    refundable = _refundable_payments()
    if not refundable:
        return jsonify({
            'success': False,
            'message': (
                'Nothing left to refund on this booking. '
                'You can still cancel with amount 0 + Cancel booking checked.'
            ),
        }), 400

    # 构建 (payment, amount) 列表
    slices = []
    if full_refund:
        slices = [(p, avail) for p, avail in refundable]
        expected_total = round(sum(a for _, a in slices), 2)
        # 前端应填 booking 可退总额；允许省略或与总额一致
        if refund_amount > 0.001 and abs(refund_amount - expected_total) > 0.02:
            return jsonify({
                'success': False,
                'message': (
                    f'Full refund amount must be ${expected_total:.2f} '
                    f'(all remaining refundable payments).'
                ),
            }), 400
        refund_amount = expected_total
    else:
        try:
            payment_id = int(payment_id_raw)
        except (TypeError, ValueError):
            return jsonify({
                'success': False,
                'message': 'Select which payment to refund (payment_id required).',
            }), 400

        match = next(((p, a) for p, a in refundable if p.id == payment_id), None)
        if not match:
            return jsonify({
                'success': False,
                'message': 'Selected payment was not found or has nothing left to refund.',
            }), 400
        payment, available = match
        refund_amount = round(refund_amount, 2)
        if refund_amount <= 0.001:
            return jsonify({'success': False, 'message': 'Invalid refund amount'}), 400
        if refund_amount > available + 0.001:
            return jsonify({
                'success': False,
                'message': (
                    f'Refund amount cannot exceed ${available:.2f} '
                    f'for the selected payment (#{payment.id}).'
                ),
            }), 400
        slices = [(payment, refund_amount)]

    def _alert_partial_refund(failed_message, *, stripe_ids_done, pairs_done, stripe_orphan_id=None):
        """Stripe/本地不一致或 Full refund 中途失败时提醒管理员。"""
        try:
            from app.ledger_alerts import send_ledger_alert
            order = booking.order_number or f'#{booking.id}'
            done_ids = [p.id for p, _ in pairs_done]
            done_amt = round(sum(a for _, a in pairs_done), 2)
            lines = [
                f'Source: refund_partial',
                f'Order: {order}',
                f'Booking id: {booking.id}',
                f'Full refund requested: {bool(full_refund)}',
                f'Manual only: {bool(manual_only)}',
                f'Committed payment_ids: {done_ids}',
                f'Committed amount: ${done_amt:.2f}',
                f'Stripe refund ids committed: {stripe_ids_done}',
                f'Stripe orphan (ledger not saved): {stripe_orphan_id or "-"}',
                f'Error: {failed_message}',
                '',
                'Do NOT blindly retry Full refund. Reconcile Stripe Dashboard vs Manage ledger first.',
            ]
            manage_url = None
            try:
                manage_url = url_for('admin.manage_trip', id=trip.id, _external=True)
            except Exception:
                manage_url = None
            send_ledger_alert(
                f'[NH Tours] Partial refund interrupt — {order}',
                '\n'.join(lines),
                alert_key=f'refund_partial:{booking.id}:{len(done_ids)}:{stripe_orphan_id or "none"}',
                headline=f'Partial refund interrupt on {order}',
                manage_url=manage_url,
            )
        except Exception as alert_err:
            current_app.logger.exception(
                'Failed to send partial-refund ledger alert booking=%s: %s',
                booking.id,
                alert_err,
            )

    def _notify_customer_refunds(pairs_done, amount_done):
        if not pairs_done:
            return
        try:
            from app.routes import send_refund_notice_email
            send_refund_notice_email(
                booking,
                pairs_done[0][0],
                amount_done,
                reason=reason,
                manual_only=manual_only,
                refunded_payments=pairs_done,
            )
        except Exception as mail_err:
            current_app.logger.exception(
                'Refund notice email raised after refund booking=%s: %s',
                booking.id,
                mail_err,
            )

    try:
        stripe_refund_ids = []
        last_ledger = None
        refunded_pairs = []
        # 每笔：行锁 →（可选）Stripe → 写账本 → 立即 commit，避免多笔中途失败时
        # Stripe 已退而本地被整单 rollback。
        for idx, (payment_ref, slice_amount) in enumerate(slices):
            is_last = idx == len(slices) - 1
            stripe_refund_id = None
            locked = (
                Payment.query.filter_by(id=payment_ref.id, booking_id=booking.id)
                .with_for_update()
                .first()
            )
            if not locked:
                db.session.rollback()
                msg = f'Payment #{payment_ref.id} disappeared during refund.'
                if refunded_pairs:
                    _alert_partial_refund(msg, stripe_ids_done=stripe_refund_ids, pairs_done=refunded_pairs)
                    done_amt = round(sum(a for _, a in refunded_pairs), 2)
                    _notify_customer_refunds(refunded_pairs, done_amt)
                    return jsonify({
                        'success': False,
                        'partial_success': True,
                        'message': (
                            f'{len(refunded_pairs)} payment(s) already refunded '
                            f'(${done_amt:.2f}), then failed: {msg}. '
                            'Do not retry Full refund blindly — check Stripe and ledger.'
                        ),
                        'stripe_refund_ids': stripe_refund_ids,
                        'payment_ids': [p.id for p, _ in refunded_pairs],
                        'new_amount_paid': float(booking.amount_paid or 0),
                        'booking_status': booking.status,
                    }), 409
                return jsonify({'success': False, 'message': msg}), 400

            available_now = payment_refundable_remaining(locked)
            if slice_amount > available_now + 0.001:
                db.session.rollback()
                msg = (
                    f'Payment #{locked.id} refundable is now ${available_now:.2f} '
                    f'(requested ${slice_amount:.2f}); another refund may have raced.'
                )
                if refunded_pairs:
                    _alert_partial_refund(msg, stripe_ids_done=stripe_refund_ids, pairs_done=refunded_pairs)
                    done_amt = round(sum(a for _, a in refunded_pairs), 2)
                    _notify_customer_refunds(refunded_pairs, done_amt)
                    return jsonify({
                        'success': False,
                        'partial_success': True,
                        'message': (
                            f'{len(refunded_pairs)} payment(s) already refunded '
                            f'(${done_amt:.2f}), then failed: {msg}'
                        ),
                        'stripe_refund_ids': stripe_refund_ids,
                        'payment_ids': [p.id for p, _ in refunded_pairs],
                        'new_amount_paid': float(booking.amount_paid or 0),
                        'booking_status': booking.status,
                    }), 409
                return jsonify({'success': False, 'message': msg}), 400

            if not manual_only:
                stripe_refund_id, err = _stripe_refund_payment(locked, slice_amount)
                if err:
                    db.session.rollback()
                    if refunded_pairs:
                        _alert_partial_refund(err, stripe_ids_done=stripe_refund_ids, pairs_done=refunded_pairs)
                        done_amt = round(sum(a for _, a in refunded_pairs), 2)
                        _notify_customer_refunds(refunded_pairs, done_amt)
                        return jsonify({
                            'success': False,
                            'partial_success': True,
                            'message': (
                                f'{len(refunded_pairs)} payment(s) already refunded '
                                f'(${done_amt:.2f}), then Stripe failed: {err}. '
                                'Do not retry Full refund blindly — reconcile first.'
                            ),
                            'stripe_refund_ids': stripe_refund_ids,
                            'payment_ids': [p.id for p, _ in refunded_pairs],
                            'new_amount_paid': float(booking.amount_paid or 0),
                            'booking_status': booking.status,
                        }), 409
                    return jsonify({'success': False, 'message': err}), 400
                if stripe_refund_id:
                    stripe_refund_ids.append(stripe_refund_id)

            try:
                last_ledger = apply_refund_to_ledger(
                    locked,
                    booking,
                    slice_amount,
                    reason=reason,
                    stripe_refund_id=stripe_refund_id,
                    cancel_booking=(cancel_booking and is_last),
                    manual_only=manual_only,
                )
                db.session.commit()
            except Exception as ledger_err:
                db.session.rollback()
                msg = str(ledger_err)
                if stripe_refund_id:
                    _alert_partial_refund(
                        f'Stripe refund {stripe_refund_id} succeeded but ledger save failed: {msg}',
                        stripe_ids_done=stripe_refund_ids[:-1],
                        pairs_done=refunded_pairs,
                        stripe_orphan_id=stripe_refund_id,
                    )
                if refunded_pairs:
                    done_amt = round(sum(a for _, a in refunded_pairs), 2)
                    _notify_customer_refunds(refunded_pairs, done_amt)
                    return jsonify({
                        'success': False,
                        'partial_success': True,
                        'message': (
                            f'{len(refunded_pairs)} payment(s) saved, then ledger failed '
                            f'after Stripe: {msg}. Reconcile before retrying.'
                        ),
                        'stripe_refund_ids': stripe_refund_ids,
                        'payment_ids': [p.id for p, _ in refunded_pairs],
                        'new_amount_paid': float(booking.amount_paid or 0),
                        'booking_status': booking.status,
                    }), 409
                return jsonify({
                    'success': False,
                    'message': (
                        f'Refund ledger failed'
                        + (f' after Stripe {stripe_refund_id}' if stripe_refund_id else '')
                        + f': {msg}'
                    ),
                }), 500

            db.session.refresh(booking)
            refunded_pairs.append((locked, slice_amount))

        done_total = round(sum(a for _, a in refunded_pairs), 2)
        _notify_customer_refunds(refunded_pairs, done_total)

        return jsonify({
            'success': True,
            'message': (
                'Full refund processed successfully'
                if full_refund and not manual_only
                else (
                    'Full manual refund recorded'
                    if full_refund
                    else (
                        'Refund processed successfully'
                        if not manual_only
                        else 'Manual refund recorded'
                    )
                )
            ),
            'stripe_refund_id': stripe_refund_ids[0] if len(stripe_refund_ids) == 1 else None,
            'stripe_refund_ids': stripe_refund_ids,
            'manual_only': manual_only,
            'full_refund': full_refund,
            'payment_id': refunded_pairs[0][0].id if len(refunded_pairs) == 1 else None,
            'payment_ids': [p.id for p, _ in refunded_pairs],
            'new_amount_paid': last_ledger['booking_amount_paid'] if last_ledger else float(booking.amount_paid or 0),
            'booking_status': last_ledger['booking_status'] if last_ledger else booking.status,
            'payment_refunded_amount': last_ledger.get('payment_refunded_amount') if last_ledger else None,
        })

    except ValueError as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error processing refund: {str(e)}")
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500


@bp.route(
    '/trips/<int:trip_id>/bookings/<int:booking_id>/participants/<int:participant_id>/withdraw',
    methods=['POST'],
)
@admin_required
def withdraw_booking_participant(trip_id, booking_id, participant_id):
    """
    将参与者标为 withdrawn（软退出）。
    JSON 可选 { restore: true } 恢复为 active。不自动退款、不改套餐数量。
    """
    from datetime import datetime

    trip = Trip.query.get_or_404(trip_id)
    booking = Booking.query.get_or_404(booking_id)
    if booking.trip_id != trip.id:
        return jsonify({'success': False, 'message': 'Booking does not belong to this trip'}), 400

    participant = BookingParticipant.query.filter_by(
        id=participant_id, booking_id=booking.id
    ).first()
    if not participant:
        return jsonify({'success': False, 'message': 'Participant not found on this booking'}), 404

    data = request.get_json(silent=True) or {}
    restore = bool(data.get('restore'))

    if restore:
        participant.status = 'active'
        participant.withdrawn_at = None
        db.session.commit()
        return jsonify({
            'success': True,
            'message': 'Participant restored to active',
            'participant_id': participant.id,
            'status': participant.status,
        })

    participant.status = 'withdrawn'
    participant.withdrawn_at = datetime.utcnow()
    db.session.commit()
    return jsonify({
        'success': True,
        'message': 'Participant marked as withdrawn',
        'participant_id': participant.id,
        'status': participant.status,
        'hint': 'Refund separately if you negotiated a refund amount.',
    })


@bp.route('/trips/<int:trip_id>/bookings/<int:booking_id>/delete', methods=['POST'])
@admin_required
def delete_booking(trip_id, booking_id):
    """删除预订"""
    from app.models import BookingAddOn, BookingPackage, InstallmentPayment, Payment
       
    trip = Trip.query.get_or_404(trip_id)
    booking = Booking.query.get_or_404(booking_id)
    
    if booking.trip_id != trip.id:
        return jsonify({'success': False, 'message': 'Booking does not belong to this trip'}), 400
    
    try:
        # 按正确顺序删除所有关联记录（避免外键约束错误）
        
        # 1. 删除所有BookingAddOn记录（直接通过booking_id查询，包括所有关联的addons）
        all_addons = BookingAddOn.query.filter_by(booking_id=booking.id).all()
        for addon in all_addons:
            db.session.delete(addon)
        
        # 2. 删除所有BookingParticipant记录
        for participant in booking.participants.all():
            db.session.delete(participant)
        
        # 3. 删除所有InstallmentPayment记录
        installments = InstallmentPayment.query.filter_by(booking_id=booking.id).all()
        for installment in installments:
            db.session.delete(installment)
        
        # 4. 删除所有Payment记录（可选：根据业务需求决定是否删除支付记录）
        # 通常建议保留支付记录用于审计，但如果是pending状态的订单，可以删除
        payments = Payment.query.filter_by(booking_id=booking.id).all()
        for payment in payments:
            # 只删除pending状态的支付记录，已支付的保留用于审计
            if payment.status == 'pending':
                db.session.delete(payment)
        
        # 5. 删除所有BookingPackage记录
        for bp in booking.booking_packages.all():
            db.session.delete(bp)
        
        # 6. 最后删除Booking本身
        db.session.delete(booking)
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Booking deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting booking: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500


@bp.route('/trips/<int:trip_id>/bookings/<int:booking_id>/send-email', methods=['POST'])
@admin_required
def send_booking_email(trip_id, booking_id):
    """给预订客户发送邮件"""
    trip = Trip.query.get_or_404(trip_id)
    booking = Booking.query.get_or_404(booking_id)
    
    # 确保 booking 属于这个 trip
    if booking.trip_id != trip.id:
        return jsonify({'success': False, 'message': 'Booking does not belong to this trip'}), 400
    
    if request.is_json:
        data = request.get_json()
    else:
        data = request.form.to_dict()
    
    recipient_email = booking.client.email
    subject = data.get('subject', f'Regarding Your Booking - {trip.title}')
    email_body = data.get('body', '')
    
    if not email_body:
        return jsonify({'success': False, 'message': 'Email body cannot be empty'}), 400
    
    # 获取发件人配置
    sender_email = current_app.config.get('SENDER_EMAIL', current_app.config.get('RECIPIENT_EMAIL', 'info@nhtours.com'))

    order_label = booking.order_number or ('#' + str(booking.id))
    client_name = (booking.client.name if booking.client else '') or 'Customer'
    contact = (current_app.config.get('REPLY_TO_EMAIL') or 'info@nhtours.com').strip()
    from app.utils import is_noreply_sender
    if is_noreply_sender(contact):
        contact = 'info@nhtours.com'
    html_body = render_branded_customer_message(
        subject_line=subject,
        brand_subtitle=trip.title or 'Booking message',
        customer_name=client_name,
        message_text=email_body,
        footer_note=f'This email is regarding your booking for: {trip.title}. Order number: {order_label}',
        show_default_signoff=True,
        contact_email=contact,
    )

    text_body = f"""
Dear {client_name},

{email_body}

Best regards,
Nexus Horizons Tours Team

---
This email is regarding your booking for: {trip.title}
Order number: {order_label}
Questions? Please email us at {contact} (this message was sent from a no-reply address).
    """
    
    # 发送邮件
    success, message = send_email_via_ses(
        sender=sender_email,
        recipient=recipient_email,
        subject=subject,
        html_body=html_body,
        text_body=text_body,
        reply_to=sender_email
    )
    
    if success:
        return jsonify({'success': True, 'message': 'Email sent successfully'})
    else:
        return jsonify({'success': False, 'message': message}), 500


@bp.route('/trips/<int:id>/messages/create', methods=['POST'])
@admin_required
def create_message(id):
    """创建或更新消息（发送、保存草稿或定时发送）"""
    from app.messaging import (
        ALLOWED_RECIPIENT_TYPES,
        deliver_message,
        forced_reply_to_email,
        get_recipients_for_trip,
    )

    trip = Trip.query.get_or_404(id)

    try:
        sender_name = request.form.get(
            'sender_name',
            current_app.config.get('SENDER_DISPLAY_NAME', 'Nexus Horizons Tours'),
        )
        reply_to_email = forced_reply_to_email()
        subject = request.form.get('subject', '')
        body_html = request.form.get('body_html', '')
        recipient_config_json = request.form.get('recipient_config', '{}')
        # 仅当显式传 status=draft 才当草稿；缺省按 send_option 发送/定时
        status = (request.form.get('status') or '').strip().lower()
        send_option = request.form.get('send_option', 'now')
        scheduled_at_str = request.form.get('scheduled_at', '')
        message_id_raw = (request.form.get('message_id') or '').strip()

        recipient_config = json.loads(recipient_config_json) if recipient_config_json else {}
        rtype = recipient_config.get('type') or 'all'
        if rtype not in ALLOWED_RECIPIENT_TYPES:
            return jsonify({'success': False, 'message': 'Invalid recipient type'}), 400
        recipient_config['type'] = rtype
        if rtype == 'package':
            if not recipient_config.get('package_id') and not recipient_config.get('addon_id'):
                return jsonify({'success': False, 'message': 'Select a package or add-on'}), 400

        recipients = get_recipients_for_trip(trip, recipient_config)
        is_draft = (status == 'draft')

        message = None
        if message_id_raw:
            try:
                mid = int(message_id_raw)
            except ValueError:
                return jsonify({'success': False, 'message': 'Invalid message id'}), 400
            message = Message.query.filter_by(id=mid, trip_id=trip.id).first()
            if not message:
                return jsonify({'success': False, 'message': 'Message not found'}), 404
            if message.status not in ('draft', 'scheduled'):
                return jsonify({'success': False, 'message': 'Only drafts/scheduled messages can be updated'}), 400

        if message is None:
            message = Message(
                trip_id=trip.id,
                sender_name=sender_name,
                reply_to_email=reply_to_email,
                subject=subject or '(No subject)',
                body_html=body_html or '',
                body_text=extract_text_from_html(body_html),
                recipient_config=recipient_config,
                total_recipients=len(recipients),
                status='draft',
                created_by_id=current_user.id if current_user else None,
            )
            db.session.add(message)
        else:
            message.sender_name = sender_name
            message.reply_to_email = reply_to_email
            message.subject = subject or '(No subject)'
            message.body_html = body_html or ''
            message.body_text = extract_text_from_html(body_html)
            message.recipient_config = recipient_config
            message.total_recipients = len(recipients)

        if is_draft:
            message.status = 'draft'
            message.scheduled_at = None
        elif send_option == 'schedule':
            if not scheduled_at_str:
                return jsonify({'success': False, 'message': 'Scheduled time is required'}), 400
            try:
                scheduled_at = None
                for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M', '%Y-%m-%dT%H:%M:%S'):
                    try:
                        scheduled_at = datetime.strptime(scheduled_at_str.strip(), fmt)
                        break
                    except ValueError:
                        continue
                if scheduled_at is None:
                    return jsonify({'success': False, 'message': 'Invalid scheduled date format'}), 400
            except Exception:
                return jsonify({'success': False, 'message': 'Invalid scheduled date format'}), 400
            if scheduled_at < datetime.utcnow() - timedelta(minutes=2):
                return jsonify({'success': False, 'message': 'Scheduled time must be after the current time'}), 400
            max_year = datetime.utcnow().year + 30
            if scheduled_at.year > max_year:
                return jsonify({'success': False, 'message': f'Scheduled year cannot be after {max_year}'}), 400
            if not recipients:
                return jsonify({'success': False, 'message': 'No recipients for this selection'}), 400
            message.scheduled_at = scheduled_at
            message.status = 'scheduled'
        else:
            subj = (subject or '').strip()
            if not subj or subj == '(No subject)':
                return jsonify({'success': False, 'message': 'Subject is required'}), 400
            if not recipients:
                return jsonify({'success': False, 'message': 'No recipients for this selection'}), 400
            message.status = 'sent'

        db.session.commit()

        if not is_draft and message.status == 'sent':
            try:
                deliver_message(message, recipients=recipients)
                db.session.commit()
            except Exception as send_err:
                db.session.rollback()
                current_app.logger.error(f'Message {message.id} deliver failed: {send_err}', exc_info=True)
                # 保留记录但标回 draft，避免假装已发送
                message = Message.query.get(message.id)
                if message:
                    message.status = 'draft'
                    db.session.commit()
                return jsonify({
                    'success': False,
                    'message': f'Email send failed: {send_err}',
                    'message_id': message.id if message else None,
                    'status': 'draft',
                }), 500

        return jsonify({
            'success': True,
            'message': 'Message saved successfully',
            'message_id': message.id,
            'status': message.status,
            'total_recipients': message.total_recipients,
            'sent_count': message.sent_count or 0,
            'failed_count': message.failed_count or 0,
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f'Error creating message: {str(e)}')
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500


@bp.route('/trips/<int:id>/messages/<int:message_id>', methods=['GET'])
@login_required
def get_message(id, message_id):
    from app.messaging import message_to_dict
    trip = Trip.query.get_or_404(id)
    message = Message.query.filter_by(id=message_id, trip_id=trip.id).first_or_404()
    return jsonify({
        'success': True,
        'message': message_to_dict(message, include_recipients=True),
    })


@bp.route('/trips/<int:id>/messages/<int:message_id>', methods=['DELETE'])
@login_required
def delete_message(id, message_id):
    trip = Trip.query.get_or_404(id)
    message = Message.query.filter_by(id=message_id, trip_id=trip.id).first_or_404()
    if message.status == 'sent':
        return jsonify({'success': False, 'message': 'Sent messages cannot be deleted'}), 400
    db.session.delete(message)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Message deleted'})


@bp.route('/trips/<int:id>/messages/<int:message_id>/cancel-schedule', methods=['POST'])
@login_required
def cancel_scheduled_message(id, message_id):
    trip = Trip.query.get_or_404(id)
    message = Message.query.filter_by(id=message_id, trip_id=trip.id).first_or_404()
    if message.status not in ('scheduled', 'sending'):
        return jsonify({'success': False, 'message': 'Message is not scheduled'}), 400
    message.status = 'draft'
    message.scheduled_at = None
    db.session.commit()
    return jsonify({'success': True, 'message': 'Schedule cancelled; saved as draft'})


def extract_text_from_html(html):
    """从HTML中提取纯文本（简单实现）"""
    text = re.sub('<[^<]+?>', '', html or '')
    text = unescape(text)
    return text.strip()
